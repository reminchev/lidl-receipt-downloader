"""Анализ на изтеглени касови бележки: история на цените, XLSX, графики,
сезонен анализ на плодове и зеленчуци."""

import json
import re
import sqlite3
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Callable, Optional

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import plotly.graph_objects as go

EUR_PER_BGN = 1.95583
EUR_INTRODUCTION_DATE = datetime(2026, 1, 1)

MONTHS_BG = {
    "януари": "01", "февруари": "02", "март": "03", "април": "04",
    "май": "05", "юни": "06", "юли": "07", "август": "08",
    "септември": "09", "октомври": "10", "ноември": "11", "декември": "12",
}

PRICE_PATTERN = r"^([А-ЯA-Z][А-ЯA-ZА-Яа-я\s\.\,\'\"\-\/\(\)0-9]+?)\s{2,}(\d+[\.,]\d{2})\s*[€BDлв#]*\s*$"
UNIT_PRICE_PATTERN = r"(\d+[\.,]\d+)\s*[xх]\s*(\d+[\.,]\d{2})"

SKIP_KEYWORDS = [
    "ОБЩА", "ОБЩО", "ПЛАТЕНО", "СУМА", "TOTAL", "PAID", "НАЛИЧНОСТ",
    "МЕЖДИННА", "ОТСТЪПКИ", "DISCOUNT", "БАНКОВА", "КАРТА",
    "ВАУЧЕР", "VOUCHER", "СДАЧА", "CHANGE", "РЕСТО", "В БРОЙ",
    "ЗА ПЛАЩАНЕ", "ПЛАЩАНЕ", "FOR PAYMENT", "PAYMENT",
    "Ном:", "Z-отчет", "Каса:", "Касиер:", "АРТИКУЛА", "Копие",
    "ЕЛ. КУПОН", "ЕЛ.КУПОН", "КУПОН",
]
SKIP_LINE_MARKERS = [
    "#Lidl Plus купон", "#Акция", "ОТСТЪПКИ",
    "МЕЖДИННА СУМА", "ОБЩА СУМА", "В БРОЙ",
    "КРЕДИТНА/ДЕБИТНА", "РЕСТО", "-----",
    "Ти спести", "#Ном:", "#Z-отчет:", "#Каса:",
]

# Падеж/тегло в края на името на продукта (за канонично име при сравнение по грамаж)
WEIGHT_SUFFIX_RE = re.compile(
    r"\s*(\d[\d\.,]*)\s*(?:КГ|KG|Г|G|МЛ|ML|Л|L|БР\.?|БР|PCS?)\s*$",
    re.IGNORECASE,
)

# Списък с плодове и зеленчуци (ключови думи, срещани в касови бележки от LIDL.bg)
FRUITS_VEGETABLES_KEYWORDS = [
    "ЯБЪЛК", "БАНАН", "ПОРТОКАЛ", "МАНДАРИН", "ЛИМОН", "ГРЕЙПФРУТ", "ГРОЗДЕ",
    "ЯГОД", "МАЛИН", "КАПИН", "КАЙСИИ", "КАЙСИЯ", "ПРАСКОВИ", "ПРАСКОВА",
    "СЛИВ", "ЧЕРЕШ", "ВИШНИ", "ВИШНА", "ПЪПЕШ", "ДИНЯ", "КРУШ", "НЕКТАРИН",
    "СМОКИН", "НАР", "БОРОВИНК", "АРОНИЯ", "КАСИС", "КЛЕМЕНТИН", "САТСУМА",
    "МАНДАРИНК", "ПОРТОКАЛОВ", "ЛИМОНОВ",
    "АВОКАДО", "МАНГО", "АНАНАС", "ПАПАЯ", "МАРАКУЯ", "КИВИ", "ЛИЧИ", "РАМБУТАН",
    "ЛОНГАН", "ДУРИАН", "ДЖАКФРУТ", "КАРАМБОЛА", "ЗВЕЗДНА", "ПИТАЯ", "ДРАКОН",
    "ГУАВА", "ТАМАРИНД", "ПОМЕЛО", "ЮЗУ", "КУМКВАТ", "ФЕЙХОА",
    "ЧЕРИМОЯ", "САПОД", "МАМЕЙ", "АCЕРОЛА", "АЦЕРОЛА", "НОНИ", "ДЕРЕН",
    "ХУРМА", "ФИНИК", "СМОКИНЯ", "ИНЖИР", "ТАМАРИЛЬ", "ФИЗАЛИС", "ГОДЖИ",
    "ОБЛЕПИХ", "ШИПК", "ДЮЛЯ", "МУШМУЛ", "ДЖАНК", "ДЖАНКА",
    "КАНИСТЕЛ", "ДЖАБУТИКАБА", "НАНГКА", "САЛАК", "СНЕЙК ФРУТ",
    "БАБАКО", "БИРИМБИ", "КАРИССА", "ЛУКУМА", "НАШПИР",
    "ЦИТРОН", "БЕРГАМОТ", "КАФИР", "ПОМПЕЛМ", "БУДДХАС ХЕНД", "ХЕНД",
    "ПЕПИНО", "КУПУАСУ", "АСАИ", "АСАЙ", "МОНСТЕРА",
    "ЯГОДОПЛОДЕН", "ЦАРИГРАДСКО", "ЦАРИГРАДСК", "АГРУС",
    "АРБУТУС", "МИРТ", "МИРТА", "МОМИНА СЪЛЗА", "РОСИЦА", "КЛЮКВА", "КЛЮКВ",
    "БРУСНИЦ", "БРУСНИКА", "МОРОШК", "МОРОШКА", "ГОЛДЕН БЕРИ", "ГОЛДЪНБЕРИ",
    "ДОМАТ", "КРАСТАВИЦ", "ЧУШК", "КАРТОФ", "МОРКОВ", "ЛУКА", "ЛУК", "ЧЕСЪН",
    "ЗЕЛЕ", "БРОКОЛ", "КАРФИОЛ", "СПАНАК", "ТИКВА", "ПАТЛАДЖАН", "ТИКВИЧК",
    "МАРУЛЯ", "РЕПИЧК", "ЦЕЛИНА", "МАГДАНОЗ", "ПРАЗ", "АСПЕРЖИ", "АРТИШОК",
    "ЦВЕКЛО", "РЯПА", "КОПЪР", "РУКОЛА", "АЙСБЕРГ", "ЕНДИВИЯ", "ЦИКОРИЯ",
    "ЦАРЕВИЦ", "ГРАХ", "БОБ", "ЛЕЩА", "ПИПЕР", "ЧИЛИ", "ХАБАНЕРО", "ЯЛАДЖА",
    "ТИКВЕН", "ЗЕЛЕН", "ЗЕЛЕНЧУК", "САЛАТ", "МАШ", "НАХУТ", "СОЯ", "ЕДАМАМЕ",
    "СЛАДКА ЦАРЕВИЦ", "БРОКОЛИ", "КЕЙЛ", "МАНГОЛД", "МАНГОЛ",
    "ОКРА", "БАМЯ", "ГОРЧИЦА", "ГОРЧИЧН", "ПАКЧОЙ", "ПАК ЧОЙ", "БОКТАЙ",
    "КИТАЙСКО ЗЕЛЕ", "НАПА", "МИЗУН", "ТАТСОЙ", "КОМАЦУН", "ШИСО",
    "ДАЙКОН", "ЯПОНСК", "РЕДИС", "ХИКАМА", "ТАРО", "ЕДОК", "ЯМ", "БАТАТ",
    "СЛАДЪК КАРТОФ", "МАНИОКА", "КАСАВ", "ЯМС", "ТОПИНАМБУР", "ЕРУСАЛИМСК",
    "ПАЩЪРНАК", "СКОРЦОНЕРА", "ЗЕЛЕНА РЕПИЧК", "ВАСАБИ", "КОЛРАБИ",
    "ФЕНЕЛ", "КОПРИНЕН", "ОРИЗОВА КАША", "БАМБУКОВ", "БАМБУК ИЗДЪНИ",
    "ЛОТОС", "ЛОТУСОВ", "ВОДНА КЕСТЕН", "СИНАПТИЧН", "СИНАП",
    "КАРАМБОЛ", "ВИТЛОФ", "ЦИКОРИЙ", "РАКОВИН", "МОРКОВИ",
    "ЗЕМНА ЯБЪЛК", "ЧЕРНА РЕПА", "АЛТАЙСК", "АЛФАЛФ", "ЛЮЦЕРН",
    "КРЕСОН", "МИКРО", "МИКРОГРИЙН", "НИКНИК", "КЪЛНОВ", "КЪЛНОВЕ",
    "БЕЙБИ СПАНАК", "БЕЙБИ РУКОЛ", "БЕЙБИ МОРКОВ", "ЧЕРРИ ДОМАТ",
    "ЧЕРИ ДОМАТ", "КОКТЕЙЛЕН ДОМАТ", "СЛИВОВ ДОМАТ",
    "ГЪБИ", "ГЪБА", "ШАМПИНИОН", "СТРИДОВА", "ШИЙТАКЕ", "ПОРТОБЕЛО",
    "МАНАТАРК", "МАНАТАРКА", "ПАЧИ", "ЛИСИЧК", "ЛИСИЧКА", "ТРЮФЕЛ",
    "МАЦУТАКЕ", "ЕРИНГИЙ", "ЕНОКИ", "КРАЛСК", "КРАЛСКА ГЪБА",
    "ДЖИНДЖИФИЛ", "КУРКУМА", "ТУРМЕРИК", "БОСИЛЕК", "МЕНТА", "РИГАН",
    "МАЩЕРКА", "РОЗМАРИН", "ЛАВАНДУЛА", "ЧУБРИЦА", "ЕСТРАГОН", "ХРЯН",
    "ДЖОДЖЕН", "КОРИАНДЪР", "КОРИАНДЪ", "ЛИМОНОВА ТРЕВА", "ПАНДАН",
    "ГАЛАНГАЛ", "ЗЕЛЕН ЛИМОН", "КИНЗА", "КИНЗОВ", "МАТОЧИН",
    "ЛАЙМ", "ЛАЙМОВ", "ЦИТРУСОВ", "ЦИТРУС",
    "ПЛОД", "ПЛОДОВ", "ПЛОДОВ СОК", "ПРЕСНИ", "СВЕЖИ", "БИО ПЛОД",
    "БИО ЗЕЛЕНЧУК", "ОРГАНИЧ", "ФЕРМЕРСК", "СЕЗОНН",
]

# Продукти, чиито имена съдържат ключова дума за плод/зеленчук, но НЕ са такива
FV_EXCLUDE_KEYWORDS = [
    "ЛУКАНК", "НАРЯЗАН", "ЯМБОЛСК", "ЯМБОЛЕН",
    "ЯХНИЯ", "ЗЕЛЕНЧУКОВА ЯХН", "БОБОТИ",
    "МАНГОЛД", "СОКОЛ", "БАЛКАНКА", "БОБИНА",
    "СЛОЕНА", "СЛИВЕНСК", "ЗЕЛЕНОСАН", "ГРАХАМ",
]

# Марки/суфикси, които не променят смисъла на продукта, но често се появяват в името.
# Те помагат да се сравняват варианти като „Кашкавал Вианга“ и „Кашкавал“ като един и същ артикул.
BRAND_SUFFIXES = (
    "ВИАНГА",
    "ВИАНГА КЪЩА",
    "ЛИДЛ",
    "LIDL",
    "SELECT",
    "PREMIUM",
    "CLASSIC",
)

# Летен сезон - месеци с по-ниски цени за свежи плодове/зеленчуци
SUMMER_MONTHS = {6, 7, 8, 9}
SUMMER_MONTHS_NAMES = "Юни, Юли, Август, Септември"
OFF_SEASON_MONTHS_NAMES = "Януари – Май, Октомври – Декември"


class ReceiptAnalyzer:
    def __init__(self, log: Optional[Callable[[str], None]] = None, db_path: Optional[str] = None):
        self.log = log or (lambda message: print(message))
        # Единица за всеки продукт: '€/кг', '€/100г' или '€' (цена за пакет)
        self.products_units = {}
        self.db_path = db_path or str(Path(__file__).with_name("lidl_local_prices.db"))
        self._conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self._conn.row_factory = sqlite3.Row
        self._init_db()

    def _init_db(self) -> None:
        """Създава локалната SQLite база данни за исторически цени по продукт и дата."""
        try:
            db_file = Path(self.db_path)
            if str(self.db_path) != ':memory:':
                db_file.parent.mkdir(parents=True, exist_ok=True)
            self._conn.execute(
                """
                CREATE TABLE IF NOT EXISTS price_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_name TEXT NOT NULL,
                    normalized_name TEXT NOT NULL,
                    date TEXT NOT NULL,
                    price REAL NOT NULL,
                    unit TEXT NOT NULL,
                    source TEXT DEFAULT 'receipt',
                    receipt_file TEXT,
                    sample_count INTEGER DEFAULT 1,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(product_name, date, unit)
                )
                """
            )
            self._conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_price_history_product_date "
                "ON price_history(normalized_name, date)"
            )
            self._conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_price_history_product_name "
                "ON price_history(product_name, date)"
            )
            self._conn.commit()
        except Exception as exc:  # pragma: no cover - safety fallback
            self.log(f"Грешка при инициализация на локалната база данни: {exc}")

    def record_price(
        self,
        product_name: str,
        date_str: str,
        price: float,
        unit: str = "€",
        source: str = "receipt",
        receipt_file: Optional[str] = None,
    ) -> None:
        """Записва цена за продукт и дата в локалната SQLite база. Ако има вече запис за същия продукт/дата/единица, се усреднява."""
        if not product_name or not date_str or price is None:
            return

        normalized_name = self.normalize_product_name(product_name)
        safe_date = date_str[:10]
        safe_unit = unit or "€"

        try:
            row = self._conn.execute(
                """
                SELECT price, sample_count
                FROM price_history
                WHERE product_name = ? AND date = ? AND unit = ?
                LIMIT 1
                """,
                (product_name, safe_date, safe_unit),
            ).fetchone()

            if row:
                old_price, sample_count = row
                new_count = int(sample_count) + 1
                avg = ((float(old_price) * int(sample_count)) + float(price)) / new_count
                self._conn.execute(
                    """
                    UPDATE price_history
                    SET price = ?, sample_count = ?, source = ?, receipt_file = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE product_name = ? AND date = ? AND unit = ?
                    """,
                    (avg, new_count, source, receipt_file, product_name, safe_date, safe_unit),
                )
            else:
                self._conn.execute(
                    """
                    INSERT INTO price_history (
                        product_name, normalized_name, date, price, unit, source, receipt_file, sample_count
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                    """,
                    (product_name, normalized_name, safe_date, float(price), safe_unit, source, receipt_file),
                )
            self._conn.commit()
        except Exception as exc:  # pragma: no cover - safety fallback
            self.log(f"Грешка при запис в локалната база данни: {exc}")

    def get_price_history(self, product_name: str) -> list:
        """Връща исторически данни за продукт от локалната база. Подава се името, без да се изисква файловете."""
        if not product_name:
            return []

        normalized_name = self.normalize_product_name(product_name)
        try:
            rows = self._conn.execute(
                """
                SELECT product_name, date, price, unit, source, receipt_file
                FROM price_history
                WHERE product_name = ? OR normalized_name = ?
                ORDER BY date ASC, created_at ASC
                """,
                (product_name, normalized_name),
            ).fetchall()
        except Exception as exc:  # pragma: no cover - safety fallback
            self.log(f"Грешка при четене от локалната база данни: {exc}")
            return []

        return [
            {
                "product_name": row[0],
                "date": row[1],
                "price": float(row[2]),
                "unit": row[3],
                "source": row[4],
                "receipt_file": row[5],
            }
            for row in rows
        ]

    def get_db_summary(self) -> dict:
        """Връща всички записани истории по продукт: {product_name: {date: price}}."""
        try:
            rows = self._conn.execute(
                """
                SELECT product_name, date, price, unit
                FROM price_history
                ORDER BY product_name ASC, date ASC
                """
            ).fetchall()
        except Exception as exc:  # pragma: no cover - safety fallback
            self.log(f"Грешка при четене от локалната база данни: {exc}")
            return {}

        summary = defaultdict(dict)
        for row in rows:
            product_name, date_str, price, unit = row
            summary[product_name][date_str] = float(price)
            self.products_units.setdefault(product_name, unit)
        return dict(summary)

    def generate_local_db_report(self, output_file: str) -> str:
        """Генерира HTML отчет от локалната база данни с графика по продукт и дата."""
        product_history = self.get_db_summary()
        if not product_history:
            raise ValueError("Локалната база данни е празна. Няма данни за анализ.")

        traces = []
        for product_name, prices_by_date in sorted(product_history.items()):
            ordered = sorted(prices_by_date.items(), key=lambda item: item[0])
            if len(ordered) < 2:
                continue
            traces.append(
                {
                    "x": [date_str for date_str, _ in ordered],
                    "y": [float(price) for _, price in ordered],
                    "mode": "lines+markers",
                    "name": product_name,
                    "line": {"width": 2},
                    "marker": {"size": 5},
                    "hovertemplate": "<b>%{fullData.name}</b><br>Дата: %{x}<br>Цена: %{y:.2f} €<extra></extra>",
                }
            )

        if not traces:
            raise ValueError("Локалната база данни няма достатъчно данни за графика.")

        html = f'''<!DOCTYPE html>
<html lang="bg">
<head>
  <meta charset="utf-8">
  <title>Lidl – Локална база данни</title>
  <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 20px; background: #f6f8fb; color: #1f2937; }}
    .card {{ background: white; border-radius: 12px; padding: 18px; box-shadow: 0 2px 10px rgba(0,0,0,0.06); }}
    h1 {{ margin-top: 0; color: #0f172a; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 20px; }}
    th, td {{ border: 1px solid #dfe7f1; padding: 8px; text-align: left; }}
    th {{ background: #e2e8f0; }}
  </style>
</head>
<body>
  <div class="card">
    <h1>📊 Локална база данни за цени</h1>
    <div id="chart" style="width:100%; height:700px;"></div>
    <table>
      <thead>
        <tr><th>Продукт</th><th>Дати</th><th>Цени</th></tr>
      </thead>
      <tbody>
        {''.join(f'<tr><td>{product_name}</td><td>{", ".join(date_str for date_str, _ in sorted(prices_by_date.items()))}</td><td>{", ".join(f"{p:.2f} €" for p in sorted(prices_by_date.values()))}</td></tr>' for product_name, prices_by_date in sorted(product_history.items()))}
      </tbody>
    </table>
  </div>
  <script>
    var traces = {json.dumps(traces, ensure_ascii=False)};
    var layout = {{
      title: 'История на цените от локалната база',
      template: 'plotly_white',
      hovermode: 'closest',
      xaxis: {{ title: 'Дата', tickangle: 30 }},
      yaxis: {{ title: 'Цена (€)' }},
      legend: {{ orientation: 'v', x: 1.02, y: 1, xanchor: 'left' }},
      height: 700,
      margin: {{ l: 60, r: 220, t: 50, b: 60 }}
    }};
    Plotly.newPlot('chart', traces, layout, {{responsive: true}});
    window.addEventListener('resize', function() {{ Plotly.Plots.resize(document.getElementById('chart')); }});
  </script>
</body>
</html>'''

        Path(output_file).write_text(html, encoding="utf-8")
        return output_file

    def parse_files(self, file_paths) -> dict:
        """Парсва множество файлове и обединява продуктите в {product: {date: price}}."""
        products_data = defaultdict(dict)
        total_receipts = 0

        for file_idx, file_path in enumerate(file_paths, 1):
            self.log(f"\nФайл {file_idx}/{len(file_paths)}: {Path(file_path).name}")
            try:
                file_products = self.parse_file(file_path)
                for product_name, dates_prices in file_products.items():
                    for date_str, price in dates_prices.items():
                        if date_str in products_data[product_name]:
                            existing = products_data[product_name][date_str]
                            products_data[product_name][date_str] = (existing + price) / 2
                        else:
                            products_data[product_name][date_str] = price

                content = Path(file_path).read_text(encoding="utf-8")
                total_receipts += content.count("БЕЛЕЖКА #")
            except Exception as e:
                self.log(f"  Грешка при четене на файл: {e}")
                continue

        self.log(f"\nОбработени {total_receipts} бележки от {len(file_paths)} файла")
        self.log(f"Намерени {len(products_data)} уникални артикула")
        return products_data

    def parse_file(self, file_path) -> dict:
        """Парсва един файл с бележки и връща {product: {date: price}}."""
        products_data = defaultdict(dict)
        content = Path(file_path).read_text(encoding="utf-8")
        receipts = content.split("БЕЛЕЖКА #")

        self.log(f"  Намерени {len(receipts) - 1} бележки за парсинг...")

        for receipt_idx, receipt in enumerate(receipts[1:], 1):
            receipt_date_str = self._parse_receipt_date(receipt)
            if receipt_date_str is None:
                self.log(f"  Пропусната бележка #{receipt_idx} - не може да се извлече дата")
                continue

            try:
                receipt_date = datetime.strptime(receipt_date_str, "%Y-%m-%d")
            except ValueError:
                self.log(f"  Пропусната бележка #{receipt_idx} - невалидна дата")
                continue

            is_bgn = "BGN" in receipt or "# лв" in receipt or "лв  #" in receipt
            is_eur = "Евро" in receipt or "# Евро #" in receipt or "EUR" in receipt

            if receipt_date < EUR_INTRODUCTION_DATE:
                conversion_rate = EUR_PER_BGN
            else:
                conversion_rate = EUR_PER_BGN if is_bgn else 1.0

            products_found = 0
            lines = receipt.split("\n")

            for i, line in enumerate(lines):
                if any(marker in line for marker in SKIP_LINE_MARKERS):
                    continue

                match = re.match(PRICE_PATTERN, line.strip())
                if not match:
                    continue

                product_name = match.group(1).strip()
                price_str = match.group(2).replace(",", ".")

                try:
                    price = float(price_str)
                except ValueError:
                    continue

                if any(keyword in product_name.upper() for keyword in SKIP_KEYWORDS):
                    continue
                if len(product_name) < 3:
                    continue
                if "x" in product_name.lower() or "х" in product_name.lower():
                    continue

                final_price = price / conversion_rate

                if i > 0:
                    unit_match = re.search(UNIT_PRICE_PATTERN, lines[i - 1].strip())
                    if unit_match:
                        unit_price = float(unit_match.group(2).replace(",", "."))
                        final_price = unit_price / conversion_rate
                        self.products_units[product_name] = "€/кг"
                    else:
                        weight_kg, unit_label = self.extract_weight_from_name(product_name)
                        if weight_kg and weight_kg > 0:
                            if unit_label == "€/100г":
                                final_price = final_price / (weight_kg * 10)
                            else:
                                final_price = final_price / weight_kg
                            self.products_units[product_name] = unit_label
                        else:
                            self.products_units.setdefault(product_name, "€")

                products_data[product_name][receipt_date_str] = final_price
                self.record_price(
                    product_name=product_name,
                    date_str=receipt_date_str,
                    price=final_price,
                    unit=self.products_units.get(product_name, "€"),
                    source="receipt",
                    receipt_file=str(file_path),
                )
                products_found += 1

            if products_found > 0:
                self.log(f"    Бележка #{receipt_idx} ({receipt_date_str}): {products_found} артикула")

        self.log(f"  От този файл: {len(products_data)} уникални артикула")
        return products_data

    def _parse_receipt_date(self, receipt: str) -> Optional[str]:
        """Извлича датата на бележката в ISO формат от различни източници."""
        header = re.search(r"Дата:\s*(\d{4})-(\d{2})-(\d{2})", receipt)
        if header:
            year, month, day = header.groups()
            return f"{year}-{month}-{day}"

        dotted = re.search(r"(\d{2})\.(\d{2})\.(\d{4})\s+\d{2}:\d{2}:\d{2}", receipt)
        if dotted:
            day, month, year = dotted.groups()
            return f"{year}-{month}-{day}"

        dotted_year_first = re.search(r"(\d{4})\.(\d{2})\.(\d{2})\s+\d{2}:\d{2}", receipt)
        if dotted_year_first:
            year, month, day = dotted_year_first.groups()
            return f"{year}-{month}-{day}"

        for month_name, month_num in MONTHS_BG.items():
            match = re.search(r"(\d{1,2})\." + month_name, receipt.lower())
            if match:
                day = match.group(1).zfill(2)
                today = date.today()
                year = today.year
                if date(year, int(month_num), int(day)) > today:
                    year -= 1
                return f"{year}-{month_num}-{day}"
        return None

    def extract_weight_from_name(self, product_name: str):
        """Извлича теглото от името на продукта. Връща (weight_kg, unit_label) или (None, None)."""
        upper = product_name.upper()
        weight_kg = None

        for pat in [r"(\d+[\.,]?\d*)\s*КГ(?!\w)", r"(\d+[\.,]?\d*)\s*KG(?![A-Z])"]:
            match = re.search(pat, upper)
            if match:
                try:
                    value = float(match.group(1).replace(",", "."))
                    if 0.05 <= value <= 25:
                        weight_kg = value
                        break
                except ValueError:
                    pass

        if weight_kg is None:
            for pat in [r"(\d+[\.,]?\d*)\s*Л(?!\w)", r"(\d+[\.,]?\d*)\s*L(?![A-Z])"]:
                match = re.search(pat, upper)
                if match:
                    try:
                        value = float(match.group(1).replace(",", "."))
                        if 0.05 <= value <= 10:
                            weight_kg = value
                            break
                    except ValueError:
                        pass

        if weight_kg is None:
            for pat in [r"(\d{2,4})\s*ГР?(?!\w)", r"(\d{2,4})\s*GR?(?![A-Z])"]:
                match = re.search(pat, upper)
                if match:
                    try:
                        value = float(match.group(1))
                        if 10 <= value <= 9999:
                            weight_kg = value / 1000.0
                            break
                    except ValueError:
                        pass

        if weight_kg is None:
            return None, None

        unit_label = "€/100г" if weight_kg < 0.5 else "€/кг"
        return weight_kg, unit_label

    def is_fruit_or_vegetable(self, product_name: str) -> bool:
        """Проверява дали продуктът е плод или зеленчук."""
        upper_name = product_name.upper()
        if any(excl in upper_name for excl in FV_EXCLUDE_KEYWORDS):
            return False
        return any(keyword in upper_name for keyword in FRUITS_VEGETABLES_KEYWORDS)

    @staticmethod
    def normalize_product_name(product_name: str) -> str:
        """Канонично име на продукта без падеж, грамаж, размер и типични маркови суфикси."""
        name = re.sub(r"\s+", " ", product_name.strip().upper()).strip(" -.")
        name = WEIGHT_SUFFIX_RE.sub("", name).strip(" -.")
        for brand in BRAND_SUFFIXES:
            brand_upper = brand.upper()
            if name.endswith(brand_upper):
                name = name[: -len(brand_upper)].strip(" -.")
                break
        return name.strip(" -.")

    @staticmethod
    def shorten_product_label(product_name: str, max_len: int = 26) -> str:
        """Съкращава дълго име за графични етикети, без да режe целия текст."""
        text = (product_name or "").strip()
        if len(text) <= max_len:
            return text
        words = text.split()
        if not words:
            return text
        compact = []
        current = ""
        for word in words:
            candidate = f"{current} {word}" if current else word
            if len(candidate) <= max_len:
                current = candidate
            else:
                if current:
                    compact.append(current)
                current = word[:max_len - 1]
                if len(current) >= max_len:
                    break
        if current:
            compact.append(current)
        shortened = " ".join(compact)
        if len(shortened) > max_len:
            shortened = shortened[: max_len - 1].rstrip() + "…"
        return shortened

    @staticmethod
    def to_per_kg_price(price: float, unit: str, weight_kg: Optional[float]) -> float:
        """Привежда цената към €/кг (за пакет с известно тегло, €/100г, или вече €/кг)."""
        if not weight_kg:
            return price
        if unit == "€/100г":
            return price * 10
        if unit == "€":
            return price / weight_kg
        return price

    def compare_years(self, products_data: dict) -> list:
        """Сравнява средни цени за 2025 срещу 2026 само на съпоставими артикули.

        Артикулите се групират по канонично име, за да се обединят варианти като
        „Кашкавал Вианга“ и „Кашкавал“, както и различни грамажи на един и същ продукт.
        Ако за един артикул има както тегловни, така и пакетни данни, се взимат само
        данните, които са сравними между годините (в повечето случаи €/кг).
        """
        yearly: dict = defaultdict(lambda: defaultdict(list))
        display_names: dict = {}
        basis_by_key: dict = {}

        grouped = defaultdict(list)
        for name, dates_prices in products_data.items():
            canonical = self.normalize_product_name(name)
            grouped[canonical].append((name, dates_prices))

        for canonical, variants in grouped.items():
            display_names[canonical] = canonical
            has_weight_data = False
            for name, _ in variants:
                weight_kg, _ = self.extract_weight_from_name(name)
                unit = self.products_units.get(name, "€/кг" if weight_kg else "€")
                if weight_kg or unit in ("€/кг", "€/100г"):
                    has_weight_data = True
                    break

            for name, dates_prices in variants:
                weight_kg, _ = self.extract_weight_from_name(name)
                unit = self.products_units.get(name, "€/кг" if weight_kg else "€")
                basis = "kg" if weight_kg or unit in ("€/кг", "€/100г") else "package"
                basis_by_key.setdefault(canonical, basis)
                if not has_weight_data and basis == "package":
                    basis_by_key[canonical] = "package"
                elif has_weight_data and basis == "kg":
                    basis_by_key[canonical] = "kg"

                for date_str, price in dates_prices.items():
                    year = date_str[:4]
                    if year not in ("2025", "2026"):
                        continue

                    if has_weight_data and basis == "package":
                        continue

                    if basis == "kg":
                        value = self.to_per_kg_price(price, unit, weight_kg)
                    else:
                        value = price
                    yearly[canonical][year].append(value)

        rows = []
        for key, years in yearly.items():
            if "2025" not in years or "2026" not in years:
                continue
            avg_2025 = sum(years["2025"]) / len(years["2025"])
            avg_2026 = sum(years["2026"]) / len(years["2026"])
            if avg_2025 <= 0:
                continue

            rows.append(
                {
                    "product": display_names.get(key, key),
                    "basis": "€/кг" if basis_by_key.get(key) == "kg" else "€/пакет",
                    "measurements_2025": len(years["2025"]),
                    "measurements_2026": len(years["2026"]),
                    "avg_2025": avg_2025,
                    "avg_2026": avg_2026,
                    "change_pct": (avg_2026 - avg_2025) / avg_2025 * 100,
                }
            )

        rows.sort(key=lambda r: abs(r["change_pct"]), reverse=True)
        return rows

    def generate_years_html(self, rows: list, output_file: str) -> None:
        """HTML отчет: дивергираща графика + таблица с търсене и сортиране."""
        total = len(rows)
        if total == 0:
            Path(output_file).write_text("<p>Няма данни.</p>", encoding="utf-8")
            return

        increased = sum(1 for r in rows if r["change_pct"] > 0)
        decreased = sum(1 for r in rows if r["change_pct"] < 0)
        avg_change = sum(r["change_pct"] for r in rows) / total
        avg_color = "#d64541" if avg_change > 0 else "#2ecc71"
        avg_sign = "+" if avg_change > 0 else ""

        top_up = max(rows, key=lambda r: r["change_pct"])
        top_dn = min(rows, key=lambda r: r["change_pct"])

        def fmt_price(p):
            return f"{p:.2f} €"

        def diff_cell(d):
            if d is None:
                return "–", "#888"
            color = "#d64541" if d > 0 else "#2ecc71"
            sign = "+" if d > 0 else ""
            return f"{sign}{d:.1f}%", color

        table_rows = ""
        for idx, row in enumerate(rows, 1):
            diff_text, diff_color = diff_cell(row["change_pct"])
            low_conf = row["measurements_2025"] == 1 and row["measurements_2026"] == 1
            opacity = "opacity:0.65;" if low_conf else ""
            row_bg = "#fff8f8" if row["change_pct"] > 0 else ("#f3fff3" if row["change_pct"] < 0 else "#f9f9f9")
            conf_badge = '<span title="Само по 1 измерване" style="color:#aaa;font-size:10px"> ⚠</span>' if low_conf else ""
            table_rows += f'''
            <tr style="background:{row_bg};{opacity}">
              <td style="text-align:center">{idx}</td>
              <td>{row["product"]}{conf_badge}</td>
              <td style="text-align:center;color:#666">{row["basis"]}</td>
              <td style="text-align:right">{fmt_price(row["avg_2025"])}</td>
              <td style="text-align:right">{fmt_price(row["avg_2026"])}</td>
              <td style="text-align:center;font-weight:bold;color:{diff_color}">{diff_text}</td>
              <td style="text-align:center;color:#999">{row["measurements_2025"]}</td>
              <td style="text-align:center;color:#999">{row["measurements_2026"]}</td>
            </tr>'''

        chart_data = [
            {
                "product": row["product"],
                "label": self.shorten_product_label(row["product"], 30),
                "basis": row["basis"],
                "avg_2025": round(row["avg_2025"], 2),
                "avg_2026": round(row["avg_2026"], 2),
                "change_pct": round(row["change_pct"], 1),
                "low_conf": row["measurements_2025"] == 1 and row["measurements_2026"] == 1,
            }
            for row in rows
        ]
        chart_json = json.dumps(chart_data, ensure_ascii=False)

        html = f'''<!DOCTYPE html>
<html lang="bg">
<head>
  <meta charset="utf-8">
  <title>Lidl – Сравнение 2025/2026</title>
  <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
  <style>
    *{{ box-sizing:border-box; }}
    body{{ font-family:'Segoe UI',Arial,sans-serif; margin:0; padding:16px 20px; background:#f4f6f9; color:#1e293b; }}
    h1{{ margin:0 0 4px; font-size:20px; color:#1e3a5f; }}
    .sub{{ color:#64748b; font-size:13px; margin-bottom:14px; }}
    .kpi-row{{ display:flex; gap:12px; flex-wrap:wrap; margin-bottom:16px; }}
    .kpi{{ background:white; border-radius:10px; padding:12px 18px; flex:1; min-width:140px;
            box-shadow:0 1px 4px rgba(0,0,0,.08); border-left:4px solid #cbd5e1; }}
    .kpi.up{{ border-color:#d64541; }}
    .kpi.dn{{ border-color:#22c55e; }}
    .kpi.neutral{{ border-color:#3b82f6; }}
    .kpi .val{{ font-size:22px; font-weight:700; line-height:1.1; }}
    .kpi .lbl{{ font-size:11px; color:#64748b; margin-top:2px; }}
    .card{{ background:white; border-radius:10px; padding:16px; box-shadow:0 1px 4px rgba(0,0,0,.08); margin-bottom:16px; }}
    .toggle-row{{ display:flex; gap:8px; align-items:center; margin-bottom:10px; flex-wrap:wrap; }}
    button{{ padding:6px 14px; border:1px solid #cbd5e1; border-radius:6px; cursor:pointer;
             background:white; color:#334155; font-size:12px; transition:all .15s; }}
    button.active, button:hover{{ background:#1e3a5f; color:white; border-color:#1e3a5f; }}
    input#search{{ padding:6px 12px; border:1px solid #cbd5e1; border-radius:6px;
                   font-size:12px; width:220px; outline:none; }}
    table{{ width:100%; border-collapse:collapse; font-size:12px; }}
    thead tr{{ background:#1e3a5f; color:white; }}
    th{{ padding:9px 8px; border:1px solid #2d5080; cursor:pointer; user-select:none; white-space:nowrap; }}
    th:hover{{ background:#2b5aa0; }}
    td{{ border:1px solid #e2e8f0; padding:7px 8px; }}
    .note{{ font-size:11px; color:#94a3b8; margin-top:8px; }}
  </style>
</head>
<body>
  <h1>Сравнение на цени Lidl: 2025 → 2026</h1>
  <div class="sub">Сортирано по абсолютна промяна &nbsp;|&nbsp; ⚠ = само по 1 измерване (по-малко надеждно)</div>

  <div class="kpi-row">
    <div class="kpi neutral">
      <div class="val">{total}</div>
      <div class="lbl">сравнени артикула</div>
    </div>
    <div class="kpi up">
      <div class="val">{avg_sign}{avg_change:.1f}%</div>
      <div class="lbl" style="color:{avg_color}">средна промяна</div>
    </div>
    <div class="kpi up">
      <div class="val">{increased}</div>
      <div class="lbl">поскъпнали</div>
    </div>
    <div class="kpi dn">
      <div class="val">{decreased}</div>
      <div class="lbl">поевтинели</div>
    </div>
    <div class="kpi up">
      <div class="val" title="{top_up["product"]}" style="font-size:14px;color:#d64541">+{top_up["change_pct"]:.1f}%</div>
      <div class="lbl">най-голямо поскъпване: {self.shorten_product_label(top_up["product"], 22)}</div>
    </div>
    <div class="kpi dn">
      <div class="val" title="{top_dn["product"]}" style="font-size:14px;color:#22c55e">{top_dn["change_pct"]:.1f}%</div>
      <div class="lbl">най-голямо поевтиняване: {self.shorten_product_label(top_dn["product"], 22)}</div>
    </div>
  </div>

  <div class="card">
    <div class="toggle-row">
      <button class="active" onclick="showChart('diverging')">% Промяна</button>
      <button onclick="showChart('grouped')">2025 vs 2026</button>
      <button id="btnTop" class="active" onclick="toggleTop()">Топ 20</button>
      <button id="btnAll" onclick="toggleTop()">Всички ({total})</button>
    </div>
    <div id="chart"></div>
  </div>

  <div class="card">
    <div class="toggle-row">
      <input id="search" type="text" placeholder="Търси артикул…" oninput="filterTable()">
      <label style="font-size:12px;color:#64748b;margin-left:8px">
        <input type="checkbox" id="hideWeak" onchange="filterTable()"> Скрий ниско-надеждни
      </label>
      <span id="count" style="margin-left:auto;font-size:12px;color:#94a3b8"></span>
    </div>
    <table id="tbl">
      <thead>
        <tr>
          <th onclick="sortTable(0)">#</th>
          <th onclick="sortTable(1)">Артикул ↕</th>
          <th onclick="sortTable(2)">Единица</th>
          <th onclick="sortTable(3)">Ср. 2025 ↕</th>
          <th onclick="sortTable(4)">Ср. 2026 ↕</th>
          <th onclick="sortTable(5)">Промяна % ↕</th>
          <th onclick="sortTable(6)">Изм. 2025</th>
          <th onclick="sortTable(7)">Изм. 2026</th>
        </tr>
      </thead>
      <tbody id="tbody">{table_rows}</tbody>
    </table>
    <div class="note">Промяната е изчислена спрямо средна цена за годината. Продукти с 1 измерване са маркирани ⚠.</div>
  </div>

  <script>
    var ALL = {chart_json};
    var showTop = true;
    var chartMode = 'diverging';

    var cfg = {{
        responsive: true, displaylogo: false,
        toImageButtonOptions: {{ format:'png', filename:'lidl_2025_2026', scale:2 }}
    }};

    function buildDiverging(data) {{
        var labels = data.map(r => r.label);
        var vals   = data.map(r => r.change_pct);
        var colors = vals.map(v => v > 0 ? (v > 20 ? '#b91c1c' : '#e87070') : (v < -20 ? '#15803d' : '#4ade80'));
        var custom = data.map(r => [r.product, r.avg_2025, r.avg_2026, r.basis]);
        return [{{
            type:'bar', orientation:'h',
            y: labels, x: vals,
            marker:{{ color: colors }},
            customdata: custom,
            hovertemplate:'<b>%{{customdata[0]}}</b><br>Промяна: <b>%{{x:+.1f}}%</b><br>2025: %{{customdata[1]:.2f}} € &nbsp; 2026: %{{customdata[2]:.2f}} €<br>Единица: %{{customdata[3]}}<extra></extra>'
        }}];
    }}

    function buildGrouped(data) {{
        var labels = data.map(r => r.label);
        var custom = data.map(r => r.product);
        return [
            {{ type:'bar', name:'2025', orientation:'h', y:labels, x:data.map(r=>r.avg_2025),
               marker:{{color:'#7fb3d3'}}, customdata:custom,
               hovertemplate:'<b>%{{customdata}}</b><br>2025: € %{{x:.2f}}<extra></extra>' }},
            {{ type:'bar', name:'2026', orientation:'h', y:labels, x:data.map(r=>r.avg_2026),
               marker:{{color:'#f0a500'}}, customdata:custom,
               hovertemplate:'<b>%{{customdata}}</b><br>2026: € %{{x:.2f}}<extra></extra>' }}
        ];
    }}

    function drawChart() {{
        var data = showTop ? ALL.slice(0,20) : ALL;
        data = data.slice().reverse();   // най-голяма промяна отгоре
        var traces, layout;

        if (chartMode === 'diverging') {{
            traces = buildDiverging(data);
            var maxAbs = Math.max(0.1, ...data.map(r => Math.abs(r.change_pct)));
            layout = {{
                title: {{ text: 'Промяна в цената (%) 2025 → 2026', x:0.5, xanchor:'center', font:{{size:15}} }},
                template:'plotly_white',
                height: Math.max(380, data.length * 34 + 180),
                xaxis: {{ title:'% промяна', zeroline:true, zerolinecolor:'#94a3b8', zerolinewidth:2,
                          gridcolor:'#e2e8f0', range:[-maxAbs*1.25, maxAbs*1.35] }},
                yaxis: {{ gridcolor:'#e2e8f0', automargin:true, tickfont:{{size:10}} }},
                margin: {{ l:20, r:60, t:60, b:50 }},
                bargap: 0.18
            }};
        }} else {{
            traces = buildGrouped(data);
            layout = {{
                title: {{ text: 'Средна цена 2025 vs 2026', x:0.5, xanchor:'center', font:{{size:15}} }},
                template:'plotly_white', barmode:'group',
                height: Math.max(380, data.length * 48 + 200),
                xaxis: {{ title:'Цена (€)', zeroline:false, gridcolor:'#e2e8f0' }},
                yaxis: {{ gridcolor:'#e2e8f0', automargin:true, tickfont:{{size:10}} }},
                margin: {{ l:20, r:40, t:60, b:50 }},
                legend: {{ orientation:'h', x:0.5, xanchor:'center', y:1.08 }},
                bargap: 0.2
            }};
        }}
        Plotly.react('chart', traces, layout, cfg);
    }}

    function showChart(mode) {{
        chartMode = mode;
        document.querySelectorAll('.toggle-row button').forEach(b => {{
            if (b.textContent.startsWith('% ') || b.textContent.startsWith('2025')) b.classList.remove('active');
        }});
        event.target.classList.add('active');
        drawChart();
    }}

    function toggleTop() {{
        showTop = !showTop;
        document.getElementById('btnTop').classList.toggle('active', showTop);
        document.getElementById('btnAll').classList.toggle('active', !showTop);
        drawChart();
    }}

    drawChart();

    // ── table filter & sort ─────────────────────────────────────────────────
    var sortDir = {{}};
    function filterTable() {{
        var q = document.getElementById('search').value.toLowerCase();
        var hideWeak = document.getElementById('hideWeak').checked;
        var rows = document.querySelectorAll('#tbody tr');
        var vis = 0;
        rows.forEach(function(tr) {{
            var txt = tr.cells[1].textContent.toLowerCase();
            var weak = tr.cells[1].textContent.includes('⚠');
            var show = txt.includes(q) && !(hideWeak && weak);
            tr.style.display = show ? '' : 'none';
            if (show) vis++;
        }});
        document.getElementById('count').textContent = vis + ' / ' + rows.length + ' артикула';
    }}
    filterTable();

    function sortTable(col) {{
        var tbody = document.getElementById('tbody');
        var rows = Array.from(tbody.querySelectorAll('tr'));
        sortDir[col] = !sortDir[col];
        rows.sort(function(a, b) {{
            var va = a.cells[col].textContent.trim();
            var vb = b.cells[col].textContent.trim();
            var na = parseFloat(va.replace(/[^\\d.\\-+]/g,''));
            var nb = parseFloat(vb.replace(/[^\\d.\\-+]/g,''));
            if (!isNaN(na) && !isNaN(nb)) return sortDir[col] ? na-nb : nb-na;
            return sortDir[col] ? va.localeCompare(vb,'bg') : vb.localeCompare(va,'bg');
        }});
        rows.forEach(r => tbody.appendChild(r));
    }}
  </script>
</body>
</html>'''

        Path(output_file).write_text(html, encoding="utf-8")
        self.log(f"Сравнение 2025/2026 запазено: {output_file}")

    def generate_xlsx(self, products_data: dict, source_file: str) -> str:
        """Генерира XLSX файл с история на цените и връща пътя до него."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Price History"

        all_dates = set()
        for dates_prices in products_data.values():
            all_dates.update(dates_prices.keys())
        sorted_dates = sorted(all_dates)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        ws["A1"] = "Артикул"
        ws["B1"] = "Единица"
        for cell in (ws["A1"], ws["B1"]):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        for idx, date_str in enumerate(sorted_dates, start=3):
            col_letter = get_column_letter(idx)
            formatted = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
            ws[f"{col_letter}1"] = formatted
            ws[f"{col_letter}1"].font = header_font
            ws[f"{col_letter}1"].fill = header_fill
            ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2
        for product_name in sorted(products_data.keys()):
            ws[f"A{row_idx}"] = product_name
            ws[f"B{row_idx}"] = self.products_units.get(product_name, "€")
            for cell in (ws[f"A{row_idx}"], ws[f"B{row_idx}"]):
                cell.alignment = Alignment(vertical="center")
            ws[f"B{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, date_str in enumerate(sorted_dates, start=3):
                if date_str in products_data[product_name]:
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.value = products_data[product_name][date_str]
                    cell.number_format = "[$€-407] #,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for col_idx in range(1, len(sorted_dates) + 3):
                ws.cell(row=row_idx, column=col_idx).border = thin_border
            row_idx += 1

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 12
        for col_idx in range(3, len(sorted_dates) + 3):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        for cell in ws[1]:
            cell.border = thin_border

        ws.freeze_panes = "C2"

        output_file = Path(source_file).with_name(Path(source_file).stem + "_price_analysis.xlsx")
        wb.save(output_file)
        self.log(f"\nXLSX файлът е създаден успешно: {output_file}")
        return str(output_file)

    def generate_chart(self, xlsx_file: str) -> Optional[str]:
        """Генерира интерактивна HTML и статична PNG графика от XLSX файла."""
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_file)
        ws = wb.active

        dates = []
        for col in range(3, ws.max_column + 1):
            date_str = ws.cell(row=1, column=col).value
            if date_str:
                try:
                    dates.append(datetime.strptime(date_str, "%d.%m.%Y"))
                except (ValueError, TypeError):
                    continue

        if not dates:
            self.log("Не са намерени дати в XLSX файла")
            return None

        products = []
        for row_idx in range(2, ws.max_row + 1):
            product_name = ws.cell(row=row_idx, column=1).value
            if not product_name:
                continue
            prices, valid_dates = [], []
            for col_idx, date in enumerate(dates, start=3):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    prices.append(float(value))
                    valid_dates.append(date)
            if len(prices) > 5:
                products.append(
                    {
                        "name": product_name,
                        "unit": ws.cell(row=row_idx, column=2).value or "€",
                        "dates": valid_dates,
                        "prices": prices,
                    }
                )

        if not products:
            self.log("Няма продукти с повече от 5 ценови записа")
            return None

        self.log(f"Намерени {len(products)} продукта с повече от 5 цени")

        palette = [
            "#2563eb", "#16a34a", "#dc2626", "#7c3aed", "#ea580c",
            "#0f766e", "#d97706", "#db2777", "#0891b2", "#65a30d",
            "#e11d48", "#475569", "#0284c7", "#f59e0b", "#10b981",
        ]
        default_visible_products = min(len(products), 12)

        fig = go.Figure()

        for idx, product in enumerate(products):
            unit = product["unit"] or self.products_units.get(product["name"], "€")
            color = palette[idx % len(palette)]
            visible_state = True if idx < default_visible_products else "legendonly"
            fig.add_trace(
                go.Scatter(
                    x=product["dates"],
                    y=product["prices"],
                    mode="lines+markers",
                    name=product["name"],
                    visible=visible_state,
                    hovertemplate=(
                        f"<b>%{{fullData.name}}</b><br>"
                        f"Дата: %{{x|%d.%m.%Y}}<br>"
                        f"Цена: %{{y:.2f}} {unit}<br>"
                        "<extra></extra>"
                    ),
                    line=dict(color=color, width=2.4),
                    marker=dict(size=6, color=color, line=dict(width=0.5, color="white")),
                )
            )

        fig.update_layout(
            title={
                "text": (
                    f"Промяна на цените на продуктите във времето<br>"
                    f"<sub>Показани: {default_visible_products}/{len(products)} продукта • Продукти с повече от 5 записа</sub>"
                ),
                "x": 0.5,
                "xanchor": "center",
                "font": {"size": 20},
            },
            paper_bgcolor="white",
            plot_bgcolor="#f8fafc",
            xaxis=dict(title=dict(text="Дата", font=dict(size=14)), tickformat="%d.%m.%Y", gridcolor="#e2e8f0"),
            yaxis=dict(title=dict(text="Цена (€/кг · €/100г · €/пакет)", font=dict(size=14)), gridcolor="#e2e8f0"),
            hovermode="closest",
            template="plotly_white",
            height=800,
            showlegend=True,
            legend=dict(
                title=dict(text="Продукти"),
                orientation="v",
                yanchor="top",
                y=1,
                xanchor="left",
                x=1.02,
                bgcolor="rgba(255, 255, 255, 0.96)",
                bordercolor="#dfe7f1",
                borderwidth=1,
                font=dict(size=10),
                tracegroupgap=8,
            ),
            margin=dict(l=60, r=350, t=110, b=60),
        )
        fig.update_xaxes(rangeslider_visible=True)

        base_name = Path(xlsx_file).with_name(Path(xlsx_file).stem)
        html_file = f"{base_name}_interactive_chart.html"

        html_content = _build_chart_html(fig, products, self.products_units)
        Path(html_file).write_text(html_content, encoding="utf-8")
        self.log(f"Интерактивна графика запазена: {html_file}")

        try:
            self._save_static_png(products, base_name)
            self.log(f"Статична PNG графика запазена: {base_name}_chart.png")
        except Exception as e:
            self.log(f"Статичната PNG графика не можа да се генерира: {e}")

        return str(html_file)

    def _save_static_png(self, products: list, base_name: str) -> None:
        """Запазва статична PNG версия на графиката."""
        try:
            plt.style.use("seaborn-v0_8-darkgrid")
        except OSError:
            pass

        fig_height = max(8, min(20, 8 + len(products) * 0.3))
        fig_static, ax = plt.subplots(figsize=(16, fig_height))

        for product in products:
            short_name = product["name"][:35] + "..." if len(product["name"]) > 35 else product["name"]
            ax.plot(product["dates"], product["prices"], marker="o", linewidth=2, markersize=5, label=short_name, alpha=0.8)

        ax.set_xlabel("Дата", fontsize=12, weight="bold")
        ax.set_ylabel("Цена (€)", fontsize=12, weight="bold")
        ax.set_title(
            f"Промяна на цените на продуктите във времето (продукти с повече от 5 записа: {len(products)})",
            fontsize=14, weight="bold", pad=20,
        )
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m.%Y"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.xticks(rotation=45, ha="right")

        if len(products) <= 15:
            ax.legend(loc="best", fontsize=8, framealpha=0.9)
        else:
            ncols = min(3, (len(products) + 9) // 10)
            ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1), fontsize=7, framealpha=0.9, ncol=ncols)

        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        fig_static.savefig(f"{base_name}_chart.png", dpi=200, bbox_inches="tight")
        plt.close(fig_static)

    def generate_seasonal_html(self, fv_data: dict, output_file: str) -> None:
        """Генерира HTML отчет със сезонно сравнение на цените на плодове/зеленчуци."""
        seasonal_stats = []

        for product_name, dates_prices in sorted(fv_data.items()):
            summer, off_season = [], []
            for date_str, price in dates_prices.items():
                try:
                    month = int(date_str[5:7])
                except (ValueError, IndexError):
                    continue
                (summer if month in SUMMER_MONTHS else off_season).append((date_str, price))

            if not summer and not off_season:
                continue

            avg_summer = sum(p for _, p in summer) / len(summer) if summer else None
            avg_off = sum(p for _, p in off_season) / len(off_season) if off_season else None
            diff_pct = ((avg_summer - avg_off) / avg_off) * 100 if avg_summer is not None and avg_off else None

            seasonal_stats.append(
                {
                    "name": product_name,
                    "summer_prices": sorted(summer),
                    "off_season_prices": sorted(off_season),
                    "avg_summer": avg_summer,
                    "avg_off": avg_off,
                    "diff_pct": diff_pct,
                    "all_prices": sorted(list(dates_prices.items())),
                }
            )

        seasonal_stats.sort(
            key=lambda x: (0 if x["diff_pct"] is not None else 1, -abs(x["diff_pct"]) if x["diff_pct"] is not None else 0)
        )

        traces = []
        for item in seasonal_stats:
            if len(item["all_prices"]) < 2:
                continue
            if item["diff_pct"] is not None:
                color = "#e74c3c" if item["diff_pct"] > 5 else ("#27ae60" if item["diff_pct"] < -5 else "#3498db")
            else:
                color = "#95a5a6"
            traces.append(
                {
                    "x": [d for d, _ in item["all_prices"]],
                    "y": [round(p, 4) for _, p in item["all_prices"]],
                    "mode": "lines+markers",
                    "name": item["name"],
                    "line": {"width": 2, "color": color},
                    "marker": {"size": 6},
                    "hovertemplate": f'<b>{item["name"]}</b><br>Дата: %{{x}}<br>Цена: %{{y:.2f}} €<extra></extra>',
                }
            )

        shapes_js = [
            {
                "type": "rect", "xref": "x", "yref": "paper",
                "x0": f"{yr}-06-01", "x1": f"{yr}-10-01",
                "y0": 0, "y1": 1,
                "fillcolor": "rgba(255,200,0,0.10)",
                "line": {"width": 0}, "layer": "below",
            }
            for yr in range(2024, 2028)
        ]

        traces_json = json.dumps(traces, ensure_ascii=False)
        shapes_json = json.dumps(shapes_js, ensure_ascii=False)

        def fmt_price(p):
            return f"{p:.2f} €" if p is not None else "–"

        def fmt_diff(d):
            if d is None:
                return "–", "#888"
            color = "#e74c3c" if d > 5 else ("#27ae60" if d < -5 else "#888")
            arrow = "↑" if d > 0 else "↓"
            return f"{arrow} {d:+.1f}%", color

        def price_details(prices):
            return ", ".join(
                f'<span title="{d}">{round(p, 2):.2f}</span>' for d, p in prices
            ) or "–"

        table_rows = ""
        for idx, item in enumerate(seasonal_stats, 1):
            diff_text, diff_color = fmt_diff(item["diff_pct"])
            row_bg = "#f8f9fa" if idx % 2 == 0 else "white"
            table_rows += f'''
            <tr style="background:{row_bg}">
              <td style="padding:8px;border:1px solid #ddd;text-align:center">{idx}</td>
              <td style="padding:8px;border:1px solid #ddd">{item["name"]}</td>
              <td style="padding:8px;border:1px solid #ddd;text-align:center">{len(item["summer_prices"])}</td>
              <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_price(item["avg_summer"])}</td>
              <td style="padding:8px;border:1px solid #ddd;text-align:center">{len(item["off_season_prices"])}</td>
              <td style="padding:8px;border:1px solid #ddd;text-align:right">{fmt_price(item["avg_off"])}</td>
              <td style="padding:8px;border:1px solid #ddd;text-align:center;font-weight:bold;color:{diff_color}">{diff_text}</td>
              <td style="padding:8px;border:1px solid #ddd;font-size:11px;color:#555">{price_details(item["summer_prices"])}</td>
              <td style="padding:8px;border:1px solid #ddd;font-size:11px;color:#555">{price_details(item["off_season_prices"])}</td>
            </tr>'''

        html = f'''<!DOCTYPE html>
<html lang="bg">
<head>
  <meta charset="utf-8">
  <title>Lidl – Сезонен анализ на плодове и зеленчуци</title>
  <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
    h1 {{ color: #2c7a2c; }}
    h2 {{ color: #333; margin-top: 30px; }}
    .legend-box {{ display: inline-block; width: 14px; height: 14px;
                   border-radius: 3px; margin-right: 6px; vertical-align: middle; }}
    .info-banner {{ background: #fff3cd; border-left: 5px solid #ffc107;
                    padding: 12px 16px; margin-bottom: 20px; border-radius: 4px; }}
    table {{ width: 100%; border-collapse: collapse; background: white; font-size: 13px; }}
    thead tr {{ background: #2c7a2c; color: white; }}
    th {{ padding: 10px; border: 1px solid #999; }}
    .controls {{ margin-bottom: 16px; }}
    input[type=text] {{ padding: 7px; width: 260px; border: 1px solid #ccc;
                        border-radius: 4px; font-size: 14px; }}
    button {{ padding: 7px 18px; margin: 4px; border: none; border-radius: 4px;
              cursor: pointer; font-size: 13px; }}
    .btn-green {{ background: #2c7a2c; color: white; }}
    .btn-gray  {{ background: #6c757d; color: white; }}
  </style>
</head>
<body>
  <h1>🌿 Lidl – Сезонен анализ на плодове и зеленчуци</h1>

  <div class="info-banner">
    <strong>Легенда:</strong>
    <span class="legend-box" style="background:#ffc107;opacity:0.5"></span> Летен сезон ({SUMMER_MONTHS_NAMES}) |&nbsp;
    <span style="color:#e74c3c;font-weight:bold">↑ Скъпи лятото</span> (разлика &gt;5%) |&nbsp;
    <span style="color:#27ae60;font-weight:bold">↓ По-евтини лятото</span> (разлика &lt;–5%) |&nbsp;
    <span style="color:#888">≈ Без съществена разлика</span><br>
    <small>Цените са в EUR. Средната лятна цена се сравнява с извън-сезонната (жълтите ленти = юни – септември).</small>
  </div>

  <h2>📈 Ценова история</h2>
  <div class="controls">
    <input type="text" id="searchInput" placeholder="Търси продукт...">
    <button class="btn-green" onclick="filterTraces()">Филтрирай</button>
    <button class="btn-gray" onclick="showAllTraces()">Всички</button>
    <button class="btn-gray" onclick="hideAllTraces()">Скрий всички</button>
  </div>
  <div id="chart"></div>

  <h2>📊 Сравнителна таблица по сезон</h2>
  <table>
    <thead>
      <tr>
        <th>#</th>
        <th>Продукт</th>
        <th>Брой цени<br>(лято)</th>
        <th>Ср. цена лято<br>({SUMMER_MONTHS_NAMES})</th>
        <th>Брой цени<br>(извън сезон)</th>
        <th>Ср. цена извън сезон<br>({OFF_SEASON_MONTHS_NAMES})</th>
        <th>Разлика</th>
        <th>Цени лято (€)</th>
        <th>Цени извън сезон (€)</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>

  <p style="margin-top:20px;color:#666;font-size:12px">
    Генерирано от Lidl Receipt Downloader • {datetime.now().strftime("%d.%m.%Y %H:%M")}
  </p>

  <script>
    var traces = {traces_json};
    var shapes = {shapes_json};
    var layout = {{
      title: 'Цени на плодове и зеленчуци (жълто = летен сезон)',
      xaxis: {{ title: 'Дата', tickformat: '%d.%m.%Y', rangeslider: {{visible: true}} }},
      yaxis: {{ title: 'Цена (€/кг)' }},
      shapes: shapes,
      hovermode: 'closest',
      template: 'plotly_white',
      height: 600,
      legend: {{ orientation: 'v', x: 1.02, y: 1 }},
      margin: {{l:60, r:260, t:80, b:60}}
    }};
    Plotly.newPlot('chart', traces, layout, {{responsive: true}});

    function filterTraces() {{
      var q = document.getElementById('searchInput').value.toLowerCase();
      if (!q) {{ showAllTraces(); return; }}
      var vis = traces.map(t => t.name.toLowerCase().includes(q));
      Plotly.restyle('chart', {{visible: vis}});
    }}
    function showAllTraces() {{
      Plotly.restyle('chart', {{visible: traces.map(() => true)}});
      document.getElementById('searchInput').value = '';
    }}
    function hideAllTraces() {{
      Plotly.restyle('chart', {{visible: traces.map(() => 'legendonly')}});
    }}
    document.getElementById('searchInput').addEventListener('keypress', function(e) {{
      if (e.key === 'Enter') filterTraces();
    }});
  </script>
</body>
</html>'''

        Path(output_file).write_text(html, encoding="utf-8")
        self.log(f"Сезонен отчет запазен: {output_file}")

    def generate_index_html(self, output_file: str, files_info: dict) -> str:
        """Генерира лендинг HTML страница с описание на приложението и линкове към отчетите."""
        html = f'''<!DOCTYPE html>
<html lang="bg">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Lidl Цени Анализ – Начало</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      font-family: 'Segoe UI', Arial, sans-serif;
      margin: 0;
      padding: 20px;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      color: #333;
      min-height: 100vh;
    }}
    .container {{
      max-width: 900px;
      margin: 0 auto;
      background: white;
      border-radius: 12px;
      box-shadow: 0 20px 60px rgba(0,0,0,0.3);
      overflow: hidden;
    }}
    .header {{
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      color: white;
      padding: 40px 30px;
      text-align: center;
    }}
    h1 {{
      margin: 0 0 10px;
      font-size: 36px;
      font-weight: 700;
    }}
    .tagline {{
      font-size: 18px;
      opacity: 0.95;
      margin: 0;
    }}
    .content {{
      padding: 40px 30px;
    }}
    .section {{
      margin-bottom: 35px;
    }}
    .section h2 {{
      color: #667eea;
      font-size: 22px;
      margin-top: 0;
      border-bottom: 3px solid #f0f0f0;
      padding-bottom: 10px;
    }}
    .features {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
    }}
    @media (max-width: 600px) {{
      .features {{ grid-template-columns: 1fr; }}
    }}
    .feature {{
      background: #f8f9ff;
      padding: 20px;
      border-radius: 8px;
      border-left: 4px solid #667eea;
    }}
    .feature h3 {{
      color: #667eea;
      margin-top: 0;
      font-size: 16px;
    }}
    .feature p {{
      margin: 8px 0 0;
      font-size: 14px;
      color: #666;
      line-height: 1.5;
    }}
    .reports {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 12px;
    }}
    .report-link {{
      display: block;
      padding: 16px 20px;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      color: white;
      text-decoration: none;
      border-radius: 8px;
      transition: transform 0.2s, box-shadow 0.2s;
      font-weight: 600;
    }}
    .report-link:hover {{
      transform: translateY(-2px);
      box-shadow: 0 10px 25px rgba(102, 126, 234, 0.4);
    }}
    .report-link:disabled {{
      opacity: 0.5;
      cursor: not-allowed;
      transform: none;
    }}
    .report-link small {{
      display: block;
      font-size: 12px;
      opacity: 0.9;
      margin-top: 4px;
    }}
    .report-icon {{
      margin-right: 10px;
      font-size: 18px;
    }}
    .info-box {{
      background: #fff3cd;
      border-left: 4px solid #ffc107;
      padding: 16px;
      border-radius: 6px;
      margin: 20px 0;
      font-size: 14px;
      line-height: 1.6;
    }}
    .footer {{
      background: #f8f9fa;
      padding: 20px 30px;
      text-align: center;
      border-top: 1px solid #e0e0e0;
      font-size: 12px;
      color: #999;
    }}
    .timestamp {{
      font-size: 12px;
      color: #999;
    }}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <h1>🛒 Lidl Цени Анализ</h1>
      <p class="tagline">Проследявайте промените в цените на продуктите във времето</p>
    </div>

    <div class="content">
      <div class="section">
        <h2>📊 Какво прави приложението?</h2>
        <p>Приложението анализира касови бележки от Lidl България и:</p>
        <div class="features">
          <div class="feature">
            <h3>📈 История на цените</h3>
            <p>Отслеждане на цените на продуктите във времето с интерактивни графики</p>
          </div>
          <div class="feature">
            <h3>📋 XLSX експорт</h3>
            <p>Экспортиране на всички данни в Excel за допълнител анализ</p>
          </div>
          <div class="feature">
            <h3>🌱 Сезонен анализ</h3>
            <p>Сравнение на цените на плодове и зеленчуци през лятото и извън сезона</p>
          </div>
          <div class="feature">
            <h3>📊 Годишно сравнение</h3>
            <p>Промяна на цен между 2025 и 2026 година с детайлна таблица</p>
          </div>
        </div>
      </div>

      <div class="section">
        <h2>📁 Генерирани отчети</h2>
        <div class="reports">
'''
        
        # Добавяме линкове към отчетите
        if files_info.get("xlsx"):
            html += f'''          <a href="{Path(files_info["xlsx"]).name}" class="report-link" title="{files_info["xlsx"]}">
            <span class="report-icon">📊</span>XLSX – История на цените<small>Excel файл със всички цени по дати</small></a>\n'''
        
        if files_info.get("chart"):
            html += f'''          <a href="{Path(files_info["chart"]).name}" class="report-link" title="{files_info["chart"]}">
            <span class="report-icon">📈</span>Интерактивна графика<small>Линейни графики с всички продукти</small></a>\n'''
        
        if files_info.get("seasonal"):
            html += f'''          <a href="{Path(files_info["seasonal"]).name}" class="report-link" title="{files_info["seasonal"]}">
            <span class="report-icon">🌱</span>Сезонен анализ<small>Сравнение на цените на плодове и зеленчуци</small></a>\n'''
        
        if files_info.get("years"):
            html += f'''          <a href="{Path(files_info["years"]).name}" class="report-link" title="{files_info["years"]}">
            <span class="report-icon">📊</span>Годишно сравнение 2025/2026<small>Промяна на цени между годините</small></a>\n'''
        
        if files_info.get("db_report"):
            html += f'''          <a href="{Path(files_info["db_report"]).name}" class="report-link" title="{files_info["db_report"]}">
            <span class="report-icon">💾</span>Локална база данни<small>История на цени от локално хранилище</small></a>\n'''

        html += '''        </div>
      </div>

      <div class="section">
        <h2>💡 Как да използвам?</h2>
        <ol style="line-height: 1.8; color: #666;">
          <li><strong>Изтеглете касова бележка</strong> – Преведете касовата бележка в текстов файл</li>
          <li><strong>Добавете файла</strong> – Селектирайте файла в приложението</li>
          <li><strong>Пуснете анализа</strong> – Приложението обработва данните и генерира отчетите</li>
          <li><strong>Преглеждайте отчетите</strong> – Отворете генерираните HTML и Excel файлове</li>
        </ol>
      </div>

      <div class="info-box">
        <strong>ℹ️ Забележка:</strong> Всеки нов анализ генерира нови файлове в папката <code>Документи</code>. 
        Проверете времето на промяна на файловете за да видите кои са най-новите.
      </div>
    </div>

    <div class="footer">
      <div class="timestamp">Генерирано: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}</div>
      <p style="margin: 10px 0 0;">Lidl Цени Анализ v1.0</p>
    </div>
  </div>
</body>
</html>'''

        Path(output_file).write_text(html, encoding="utf-8")
        self.log(f"Лендинг страница запазена: {output_file}")
        return str(output_file)


def _build_chart_html(fig, products: list, products_units: dict) -> str:
    """Изгражда HTML-а на интерактивната графика с филтър и топ-10 таблици."""
    changes = []
    for product in products:
        prices = product["prices"]
        if len(prices) < 2 or prices[0] <= 0:
            continue
        min_i, max_i = prices.index(min(prices)), prices.index(max(prices))
        changes.append(
            {
                "name": product["name"],
                "unit": product.get("unit") or products_units.get(product["name"], "€"),
                "change_percent": ((prices[-1] - prices[0]) / prices[0]) * 100,
                "min_price": min(prices),
                "max_price": max(prices),
                "min_price_date": product["dates"][min_i].strftime("%d.%m.%Y"),
                "max_price_date": product["dates"][max_i].strftime("%d.%m.%Y"),
            }
        )

    changes.sort(key=lambda x: abs(x["change_percent"]), reverse=True)
    top_up = changes[:10]
    top_down = sorted([c for c in changes if c["change_percent"] < 0], key=lambda x: x["change_percent"])[:10]

    def table_block(items, title, header_color):
        rows = ""
        for idx, item in enumerate(items, 1):
            row_bg = "#f8f9fa" if idx % 2 == 0 else "white"
            color = "#dc3545" if item["change_percent"] > 0 else "#28a745"
            arrow = "↑" if item["change_percent"] > 0 else "↓"
            rows += f'''
            <tr style="background-color: {row_bg};">
                <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{idx}</td>
                <td style="padding: 10px; border: 1px solid #ddd;">{item["name"]} <small style="color:#888">({item["unit"]})</small></td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: {color};">{arrow} {item["change_percent"]:+.2f}%</td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">{item["min_price"]:.2f}<br><small style="color: #666;">({item["min_price_date"]})</small></td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">{item["max_price"]:.2f}<br><small style="color: #666;">({item["max_price_date"]})</small></td>
            </tr>'''
        return f'''
            <div class="table-container" style="margin-top: 30px; padding: 20px; background-color: white; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h2 style="color: #333; margin-bottom: 20px;">{title}</h2>
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <thead>
                        <tr style="background-color: {header_color}; color: white;">
                            <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">#</th>
                            <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">Продукт</th>
                            <th style="padding: 12px; text-align: center; border: 1px solid #ddd;">Промяна (%)</th>
                            <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Мин. цена (€)</th>
                            <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Макс. цена (€)</th>
                        </tr>
                    </thead>
                    <tbody>{rows}</tbody>
                </table>
            </div>'''

    top_up_html = table_block(top_up, "📊 Топ 10 продукти с най-голяма ценова промяна", "#007bff")
    top_down_html = table_block(top_down, "📉 Топ 10 продукти с най-голямо понижение на цените", "#28a745") if top_down else ""

    return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Lidl - Интерактивна графика на цените</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 100%;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .controls {{ margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; }}
        .control-group {{ margin-bottom: 15px; }}
        label {{ font-weight: bold; margin-right: 10px; display: inline-block; width: 150px; }}
        input[type="text"] {{ padding: 8px; width: 300px; border: 1px solid #ddd; border-radius: 4px; }}
        button {{ padding: 8px 20px; margin: 5px; border: none; border-radius: 4px; cursor: pointer; font-size: 14px; }}
        .btn-primary {{ background-color: #007bff; color: white; }}
        .btn-primary:hover {{ background-color: #0056b3; }}
        .btn-secondary {{ background-color: #6c757d; color: white; }}
        .btn-secondary:hover {{ background-color: #545b62; }}
        .btn-success {{ background-color: #28a745; color: white; }}
        .btn-success:hover {{ background-color: #218838; }}
        .info {{ margin-top: 10px; padding: 10px; background-color: #d1ecf1; border-left: 4px solid #0c5460; color: #0c5460; }}
        #chart {{ margin-top: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🛒 Lidl - Интерактивна графика на цените</h1>

        <div class="controls">
            <div class="control-group">
                <label>🔍 Търси продукт:</label>
                <input type="text" id="searchInput" placeholder="Въведете име на продукт...">
                <button class="btn-primary" onclick="filterProducts()">Филтрирай</button>
                <button class="btn-success" onclick="showTopProducts()">Покажи първите 12</button>
            </div>

            <div class="control-group">
                <button class="btn-success" onclick="showAll()">Покажи всички</button>
                <button class="btn-secondary" onclick="hideAll()">Скрий всички</button>
                <button class="btn-secondary" onclick="resetView()">Възстанови изглед</button>
            </div>

            <div class="info">
                <strong>💡 Съвети:</strong>
                <ul style="margin: 5px 0;">
                    <li>По подразбиране показваме първите 12 продукта, за да не се губите при голям брой линии</li>
                    <li>Кликнете на продукт в легендата за да го покажете/скриете</li>
                    <li>Използвайте мишката за приближаване (scroll) и местене (drag)</li>
                    <li>Използвайте филтъра за търсене на конкретни продукти</li>
                    <li>Двоен клик на легендата изолира един продукт</li>
                </ul>
            </div>
        </div>

        <div id="chart"></div>

        {top_up_html}

        {top_down_html}
    </div>

    <script>
        var plotData = {fig.to_json()};
        var layout = plotData.layout;
        var data = plotData.data;
        var config = {{
            responsive: true,
            displayModeBar: true,
            modeBarButtonsToAdd: ['drawopenpath', 'eraseshape'],
            toImageButtonOptions: {{
                format: 'png',
                filename: 'lidl_prices_chart',
                height: 1080,
                width: 1920,
                scale: 2
            }}
        }};

        var originalVisibility = data.map(trace => trace.visible);
        var defaultVisibleCount = Math.min(12, data.length);
        var initialVisibility = data.map((trace, index) => index < defaultVisibleCount);

        if (!layout.legend) {{
            layout.legend = {{}};
        }}
        layout.legend.title = {{text: 'Продукти'}};
        layout.legend.font = {{size: 10}};
        layout.legend.tracegroupgap = 8;

        Plotly.newPlot('chart', data, layout, config);
        Plotly.restyle('chart', {{visible: initialVisibility}});

        function filterProducts() {{
            var searchText = document.getElementById('searchInput').value.toLowerCase();
            if (!searchText) {{ showTopProducts(); return; }}
            Plotly.restyle('chart', {{
                visible: data.map(function(trace) {{
                    return trace.name.toLowerCase().includes(searchText);
                }})
            }});
        }}

        function showTopProducts() {{
            Plotly.restyle('chart', {{visible: data.map((trace, index) => index < defaultVisibleCount)}});
            document.getElementById('searchInput').value = '';
        }}

        function showAll() {{
            Plotly.restyle('chart', {{visible: data.map(() => true)}});
            document.getElementById('searchInput').value = '';
        }}

        function hideAll() {{
            Plotly.restyle('chart', {{visible: data.map(() => 'legendonly')}});
        }}

        function resetView() {{
            Plotly.relayout('chart', {{
                'xaxis.autorange': true,
                'yaxis.autorange': true
            }});
        }}

        document.getElementById('searchInput').addEventListener('keypress', function(e) {{
            if (e.key === 'Enter') {{ filterProducts(); }}
        }});
    </script>
</body>
</html>'''