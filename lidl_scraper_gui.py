"""
Lidl Receipt Downloader - GUI Version
Автоматично изтегля всички касови бележки от Lidl.bg с филтриране по дата
"""

import asyncio
import os
import threading
import time
import re
import json
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
from tkcalendar import DateEntry
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import plotly.graph_objects as go

class LidlReceiptDownloader:
    def __init__(self, output_dir: str, start_date=None, end_date=None, log_callback=None, progress_callback=None):
        self.output_dir = output_dir
        self.start_date = start_date
        self.end_date = end_date
        self.receipts = []
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self.is_cancelled = False
        self.ready_to_start = False
        self.total_pages_estimated = 100  # Оценка - 10 бележки на страница
        self.current_page_processed = 0
        self.start_time = None
        
    def log(self, message):
        """Логване на съобщение"""
        if self.log_callback:
            self.log_callback(message)
        print(message)
    
    def parse_receipt_date(self, text_content):
        """Извлича датата от касовата бележка"""
        try:
            # Търсим дата във формат DD.MM.YYYY HH:MM:SS
            date_pattern = r'(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2}:\d{2})'
            match = re.search(date_pattern, text_content)
            if match:
                date_str = match.group(1)
                # Конвертираме във формат YYYY-MM-DD за сравнение
                parts = date_str.split('.')
                return f"{parts[2]}-{parts[1]}-{parts[0]}"
        except (AttributeError, IndexError, ValueError):
            pass
        return None
    
    def is_date_in_range(self, receipt_date_str):
        """Проверява дали датата е в избрания период"""
        if not self.start_date and not self.end_date:
            return True
        
        if not receipt_date_str:
            return True  # Ако не можем да извлечем дата, включваме бележката
        
        try:
            if self.start_date and receipt_date_str < self.start_date:
                return False
            if self.end_date and receipt_date_str > self.end_date:
                return False
            return True
        except:
            return True
    
    async def wait_for_user_ready(self, page):
        """Изчаква потребителя да се позиционира на страницата с касови бележки"""
        self.log("📌 ИНСТРУКЦИИ:")
        self.log("=" * 60)
        self.log("1. Влезте в акаунта си в отворения браузър")
        self.log("2. Отидете на страницата с касови бележки")
        self.log("   (https://www.lidl.bg/mre/purchase-history)")
        self.log("3. Натиснете 'Започни изтегляне' когато сте готови")
        self.log("=" * 60)
        
        try:
            await page.goto('https://accounts.lidl.com/Account/Login?ReturnUrl=%2Fconnect%2Fauthorize%2Fcallback%3Fcountry_code%3DBG%26response_type%3Dcode%26client_id%3Dbulgariaretailclient%26scope%3Dopenid%2520profile%2520Lidl.Authentication%2520offline_access%26state%3D7kjyF6Xd4NaMWmVqiNhXmDlKvTzcOa23tPkuFORkF2E%253D%26redirect_uri%3Dhttps%253A%252F%252Fwww.lidl.bg%252Fuser-api%252Fsignin-oidc%26nonce%3DEJIGNwoTYnT5BTScAf8yndJ6_tfF5V-ag26aqBsTg-8%26step%3Dlogin%26language%3Dbg-BG#login', wait_until='networkidle')
            
            self.log("\n⏳ Изчакване на потребителя...")
            self.log("   Моля, влезте и отидете на страницата с касови бележки\n")
            
            # Изчакваме сигнал за продължаване
            while not self.ready_to_start and not self.is_cancelled:
                await asyncio.sleep(0.5)
            
            if self.is_cancelled:
                return
            
            self.log("✓ Стартиране на изтегляне на бележки...")
            self.start_time = time.time()
            await asyncio.sleep(1)
            
        except Exception as e:
            self.log(f"❌ Грешка: {e}")
            raise
    
    async def navigate_to_purchase_history(self, page, page_num=1):
        """Отива до страницата с покупки"""
        url = f'https://www.lidl.bg/mre/purchase-history?client_id=BulgariaRetailClient&country_code=bg&language=bg-BG&page={page_num}'
        self.log(f"Отваряне на история на покупките (страница {page_num})...")
        await page.goto(url, wait_until='networkidle')
        await asyncio.sleep(2)
    
    async def extract_receipts_from_page(self, page, page_number):
        """Извлича касовите бележки от текущата страница"""
        self.log("Извличане на касови бележки...")
        
        try:
            await asyncio.sleep(3)
            
            purchase_selectors = [
                'a[href*="/mre/purchase-detail"]',
                'a.card[href*="purchase-detail"]',
                'a[data-testid][class*="card"]',
                'a.card',
                'a[class*="card"][href*="/mre/"]'
            ]
            
            purchase_elements = []
            for selector in purchase_selectors:
                elements = await page.query_selector_all(selector)
                if elements:
                    purchase_elements = elements
                    self.log(f"Използван селектор: {selector}")
                    break
            
            if not purchase_elements:
                purchase_elements = await page.query_selector_all('button, a[href*="purchase"], a[href*="receipt"]')
            
            self.log(f"Намерени {len(purchase_elements)} покупки на тази страница")
            
            # Предполагаме 10 бележки на страница, коригираме оценката
            if len(purchase_elements) > 0:
                self.total_pages_estimated = max(self.total_pages_estimated, page_number + 10)
            
            for i in range(len(purchase_elements)):
                if self.is_cancelled:
                    self.log("⚠ Процесът е прекъснат от потребителя")
                    return
                
                try:
                    self.log(f"  Обработка на покупка {i + 1}/{len(purchase_elements)}...")
                    
                    await asyncio.sleep(1)
                    used_selector = 'a.card[href*="purchase-detail"]'
                    for selector in purchase_selectors:
                        test_elements = await page.query_selector_all(selector)
                        if test_elements:
                            used_selector = selector
                            break
                    current_elements = await page.query_selector_all(used_selector)
                    
                    if i >= len(current_elements):
                        self.log(f"  Елемент {i + 1} вече не е достъпен, прескачане...")
                        continue
                    
                    element = current_elements[i]
                    
                    await element.scroll_into_view_if_needed()
                    await asyncio.sleep(0.5)
                    
                    await element.click()
                    await asyncio.sleep(2)
                    
                    await page.wait_for_load_state('networkidle', timeout=10000)
                    await asyncio.sleep(2)
                    
                    # Изчакване за зареждане на бележката
                    await asyncio.sleep(1)
                    
                    receipt_selectors = [
                        'main',
                        'body'
                    ]
                    
                    text_content = None
                    for selector in receipt_selectors:
                        receipt_container = await page.query_selector(selector)
                        if receipt_container:
                            text_content = await receipt_container.inner_text()
                            if text_content and len(text_content.strip()) > 100:
                                break
                    
                    if not text_content:
                        text_content = await page.inner_text('body')
                    
                    # Почистване на ненужен текст от navigation и footer
                    if text_content:
                        # Извличаме само основното съдържание на бележката
                        lines = text_content.split('\n')
                        # Премахваме първите редове ако са navigation
                        cleaned_lines = []
                        start_found = False
                        for line in lines:
                            # Търсим начало на бележката (обикновено със БУЛСТАТ или адрес на магазин)
                            if 'БУЛСТАТ' in line or 'УНП' in line or 'Лидл' in line or not start_found:
                                start_found = True
                            if start_found:
                                cleaned_lines.append(line)
                        text_content = '\n'.join(cleaned_lines).strip()
                    
                    if text_content and text_content.strip():
                        # Проверка на датата
                        receipt_date = self.parse_receipt_date(text_content)
                        
                        if self.is_date_in_range(receipt_date):
                            receipt_data = {
                                'page_number': page_number,
                                'index': i + 1,
                                'date': receipt_date,
                                'content': text_content.strip()
                            }
                            self.receipts.append(receipt_data)
                            date_info = f" ({receipt_date})" if receipt_date else ""
                            self.log(f"    ✓ Извлечена бележка {i + 1}{date_info} ({len(text_content)} символа)")
                            self.log(f"    📊 Общо изтеглени бележки: {len(self.receipts)}")
                        else:
                            date_info = f" ({receipt_date})" if receipt_date else ""
                            self.log(f"    ⊗ Пропусната бележка {i + 1}{date_info} - извън период")
                    else:
                        self.log(f"    ⚠ Бележка {i + 1} е празна")
                    
                    await page.go_back()
                    await asyncio.sleep(2)
                    await page.wait_for_load_state('networkidle', timeout=10000)
                    
                except PlaywrightTimeout:
                    self.log(f"  Timeout при обработка на покупка {i + 1}, продължаване...")
                    try:
                        await page.go_back()
                        await asyncio.sleep(2)
                    except:
                        pass
                except Exception as e:
                    self.log(f"  Грешка при обработка на покупка {i + 1}: {e}")
                    try:
                        await page.go_back()
                        await asyncio.sleep(2)
                    except:
                        pass
                    continue
                    
        except Exception as e:
            self.log(f"Грешка при извличане на бележки: {e}")
    
    async def has_more_receipts(self, page):
        """Проверява дали има още бележки (покупки) на страницата"""
        try:
            # Проверка за покупки на страницата
            purchase_links = await page.query_selector_all('a[href*="/mre/purchase-detail"]')
            return len(purchase_links) > 0
        except Exception:
            return False
    
    async def check_current_page_number(self, page):
        """Извлича текущия номер на страницата от URL"""
        try:
            current_url = page.url
            if 'page=' in current_url:
                page_param = current_url.split('page=')[1].split('&')[0]
                return int(page_param)
            return 1
        except:
            return 1
    
    async def download_all_receipts(self):
        """Изтегля всички касови бележки"""
        async with async_playwright() as p:
            self.log("Стартиране на браузър...")
            
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            )
            page = await context.new_page()
            
            try:
                await self.wait_for_user_ready(page)
                
                # Получаваме текущия URL за да определим началната страница
                current_url = page.url
                page_number = await self.check_current_page_number(page)
                
                while not self.is_cancelled:
                    self.log(f"\n{'=' * 60}")
                    self.log(f"СТРАНИЦА {page_number}")
                    self.log(f"{'=' * 60}")
                    
                    # Отваряне на страницата с покупки
                    await self.navigate_to_purchase_history(page, page_number)
                    
                    # Проверка дали има покупки на страницата
                    if not await self.has_more_receipts(page):
                        self.log(f"\n✓ Няма повече покупки на страница {page_number}")
                        break
                    
                    receipts_before = len(self.receipts)
                    await self.extract_receipts_from_page(page, page_number)
                    receipts_after = len(self.receipts)
                    
                    self.current_page_processed = page_number
                    if self.progress_callback:
                        progress_percent = min(100, (page_number / self.total_pages_estimated) * 100)
                        elapsed_time = time.time() - self.start_time if self.start_time else 0
                        self.progress_callback(progress_percent, page_number, self.total_pages_estimated, elapsed_time)
                    
                    if self.is_cancelled:
                        self.log("\n⚠ Процесът е прекъснат")
                        break
                    
                    self.log(f"\n📋 Изтеглени от тази страница: {receipts_after - receipts_before}")
                    self.log(f"📊 Общо изтеглени бележки: {receipts_after}")
                    
                    # Преминаваме към следващата страница
                    page_number += 1
                
                if not self.is_cancelled:
                    self.log(f"\n{'=' * 60}")
                    self.log(f"✓ ПРИКЛЮЧЕНО ИЗТЕГЛЯНЕ")
                    self.log(f"{'=' * 60}")
                    self.log(f"📊 Общо извлечени бележки: {len(self.receipts)}")
                    elapsed_time = time.time() - self.start_time if self.start_time else 0
                    self.log(f"⏱ Общо време: {self.format_time(elapsed_time)}")
                    self.log(f"{'=' * 60}")
                
            except Exception as e:
                self.log(f"❌ Грешка при изтегляне: {e}")
                raise
                
            finally:
                await browser.close()
    
    def format_time(self, seconds):
        """Форматира времето в часове:минути:секунди"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        if hours > 0:
            return f"{hours}ч {minutes}м {secs}с"
        elif minutes > 0:
            return f"{minutes}м {secs}с"
        else:
            return f"{secs}с"
    
    def save_to_file(self):
        """Запазва бележките във файл"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'lidl_receipts_{timestamp}.txt'
        filepath = os.path.join(self.output_dir, filename)
        
        full_path = os.path.abspath(filepath)
        
        self.log(f"\n{'=' * 80}")
        self.log(f"📝 ЗАПАЗВАНЕ НА БЕЛЕЖКИ")
        self.log(f"{'=' * 80}")
        self.log(f"Брой бележки: {len(self.receipts)}")
        self.log(f"Файл: {filename}")
        self.log(f"Директория: {self.output_dir}")
        self.log(f"Пълен път: {full_path}")
        self.log(f"{'=' * 80}")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("КАСОВИ БЕЛЕЖКИ ОТ LIDL.BG\n")
            f.write(f"Дата на изтегляне: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
            f.write(f"Общо бележки: {len(self.receipts)}\n")
            if self.start_date or self.end_date:
                f.write(f"Период: ")
                if self.start_date:
                    f.write(f"от {self.start_date} ")
                if self.end_date:
                    f.write(f"до {self.end_date}")
                f.write("\n")
            f.write("=" * 80 + "\n\n")
            
            for i, receipt in enumerate(self.receipts, 1):
                f.write(f"\n{'=' * 80}\n")
                f.write(f"БЕЛЕЖКА #{i}\n")
                f.write(f"Страница: {receipt['page_number']}\n")
                if receipt.get('date'):
                    f.write(f"Дата: {receipt['date']}\n")
                f.write(f"{'=' * 80}\n\n")
                f.write(receipt['content'])
                f.write("\n\n")
        
        file_size = os.path.getsize(full_path) / 1024  # KB
        
        self.log(f"\n{'=' * 80}")
        self.log(f"✅ УСПЕШНО ЗАВЪРШЕНО!")
        self.log(f"{'=' * 80}")
        self.log(f"📊 Общо бележки: {len(self.receipts)}")
        self.log(f"📁 Файл: {filename}")
        self.log(f"📂 Пълен път: {full_path}")
        self.log(f"💾 Размер: {file_size:.2f} KB")
        self.log(f"{'=' * 60}")
        
        return full_path


class LidlGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Lidl Receipt Downloader")
        self.root.geometry("800x850")
        self.root.resizable(True, True)
        
        self.downloader = None
        self.download_thread = None
        self.output_dir = str(Path.home() / "Documents")
        self.analysis_files = []
        self.config_file = "config.json"
        
        # Зареждане на конфигурацията
        self.load_config()
        
        self.setup_ui()
        
        # Зареждане на запазения файл за анализ след създаване на UI
        self.load_saved_analysis_file()
        
    def setup_ui(self):
        """Създава интерфейса"""
        # Заглавие
        title_frame = ttk.Frame(self.root, padding="10")
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        title_label = ttk.Label(
            title_frame, 
            text="Lidl Receipt Downloader", 
            font=("Arial", 16, "bold")
        )
        title_label.grid(row=0, column=0, pady=5)
        
        subtitle_label = ttk.Label(
            title_frame, 
            text="Автоматично изтегляне на касови бележки от Lidl.bg",
            font=("Arial", 10)
        )
        subtitle_label.grid(row=1, column=0, pady=2)
        
        # Рамка за период
        period_frame = ttk.LabelFrame(self.root, text="СТЪПКА 1: Период на бележките (опционално)", padding="10")
        period_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        ttk.Label(period_frame, text="От дата:").grid(row=0, column=0, sticky=tk.W, pady=5, padx=5)
        self.start_date_entry = DateEntry(
            period_frame,
            width=18,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            state='normal'
        )
        self.start_date_entry.grid(row=0, column=1, sticky=(tk.W), pady=5, padx=5)
        
        # Бутон за изчистване на начална дата
        self.clear_start_btn = ttk.Button(
            period_frame,
            text="✖",
            width=3,
            command=self.clear_start_date
        )
        self.clear_start_btn.grid(row=0, column=2, pady=5, padx=2)
        
        ttk.Label(period_frame, text="До дата:").grid(row=1, column=0, sticky=tk.W, pady=5, padx=5)
        self.end_date_entry = DateEntry(
            period_frame,
            width=18,
            background='darkblue',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            state='normal'
        )
        self.end_date_entry.grid(row=1, column=1, sticky=(tk.W), pady=5, padx=5)
        
        # Бутон за изчистване на крайна дата
        self.clear_end_btn = ttk.Button(
            period_frame,
            text="✖",
            width=3,
            command=self.clear_end_date
        )
        self.clear_end_btn.grid(row=1, column=2, pady=5, padx=2)
        
        # Checkbox за използване на период
        self.use_period_var = tk.BooleanVar(value=False)
        self.use_period_check = ttk.Checkbutton(
            period_frame,
            text="✓ Филтрирай по период",
            variable=self.use_period_var
        )
        self.use_period_check.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=5, padx=5)
        
        period_frame.columnconfigure(3, weight=1)
        
        # Рамка за директория
        dir_frame = ttk.LabelFrame(self.root, text="СТЪПКА 2: Директория за съхранение", padding="10")
        dir_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        self.dir_label = ttk.Label(dir_frame, text=self.output_dir, foreground="blue")
        self.dir_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        self.dir_button = ttk.Button(
            dir_frame, 
            text="📁 Избери директория", 
            command=self.choose_directory
        )
        self.dir_button.grid(row=0, column=1, padx=5)
        
        dir_frame.columnconfigure(0, weight=1)
        
        # Рамка за контроли на изтегляне
        download_control_frame = ttk.LabelFrame(self.root, text="СТЪПКА 3: Изтегляне на бележки", padding="10")
        download_control_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        # Описание
        desc_label = ttk.Label(
            download_control_frame,
            text="Натиснете 'Старт', влезте в акаунта си в отворения браузър, после натиснете 'Започни изтегляне'",
            font=("Arial", 9),
            foreground="darkblue"
        )
        desc_label.grid(row=0, column=0, columnspan=5, pady=(0, 10), sticky=tk.W)
        
        self.start_button = ttk.Button(
            download_control_frame, 
            text="① Старт", 
            command=self.start_download,
            style="Accent.TButton",
            width=20
        )
        self.start_button.grid(row=1, column=0, padx=5, pady=5)
        
        arrow_label1 = ttk.Label(download_control_frame, text="➜", font=("Arial", 14))
        arrow_label1.grid(row=1, column=1, padx=5)
        
        self.continue_button = ttk.Button(
            download_control_frame, 
            text="② Започни изтегляне", 
            command=self.continue_after_ready,
            state=tk.DISABLED,
            style="Accent.TButton",
            width=20
        )
        self.continue_button.grid(row=1, column=2, padx=5, pady=5)
        
        arrow_label2 = ttk.Label(download_control_frame, text="➜", font=("Arial", 14))
        arrow_label2.grid(row=1, column=3, padx=5)
        
        self.stop_button = ttk.Button(
            download_control_frame, 
            text="⏸ Прекъсване", 
            command=self.stop_download,
            state=tk.DISABLED,
            width=20
        )
        self.stop_button.grid(row=1, column=4, padx=5, pady=5)
        
        download_control_frame.columnconfigure(5, weight=1)
        
        # Рамка за файлове за анализ
        analysis_frame = ttk.LabelFrame(self.root, text="СТЪПКА 4: Анализ на цени (опционално)", padding="10")
        analysis_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        self.analysis_file_label = ttk.Label(
            analysis_frame, 
            text="Няма избрани файлове", 
            foreground="gray"
        )
        self.analysis_file_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        self.analysis_file_button = ttk.Button(
            analysis_frame, 
            text="📄 Избери файлове за анализ", 
            command=self.choose_analysis_files
        )
        self.analysis_file_button.grid(row=0, column=1, padx=5)
        
        self.analysis_folder_button = ttk.Button(
            analysis_frame, 
            text="📁 Избери папка", 
            command=self.choose_analysis_folder
        )
        self.analysis_folder_button.grid(row=0, column=2, padx=5)
        
        # Бутон за анализ
        self.analyze_button = ttk.Button(
            analysis_frame, 
            text="📊 Анализ → XLSX", 
            command=self.analyze_receipts,
            style="Accent.TButton"
        )
        self.analyze_button.grid(row=1, column=0, columnspan=3, pady=10)
        
        analysis_frame.columnconfigure(0, weight=1)
        
        # Рамка за статус
        status_frame = ttk.Frame(self.root, padding="10")
        status_frame.grid(row=5, column=0, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        ttk.Label(status_frame, text="Статус:", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=5)
        
        # Статус лейбъл
        self.status_label = ttk.Label(
            status_frame, 
            text="Готов за стартиране",
            foreground="green",
            font=("Arial", 10, "bold")
        )
        self.status_label.grid(row=0, column=1, padx=5)
        
        # Таймер лейбъл
        self.timer_label = ttk.Label(
            status_frame, 
            text="⏱ Време: 0с",
            foreground="blue",
            font=("Arial", 10)
        )
        self.timer_label.grid(row=0, column=2, padx=15)
        
        status_frame.columnconfigure(3, weight=1)
        
        # Рамка за прогрес барове
        progress_frame = ttk.LabelFrame(self.root, text="Прогрес на изтегляне", padding="10")
        progress_frame.grid(row=6, column=0, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        # Първи прогрес бар - страници
        ttk.Label(progress_frame, text="Прогрес по страници:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.page_progress = ttk.Progressbar(
            progress_frame, 
            mode='determinate',
            length=400
        )
        self.page_progress.grid(row=0, column=1, padx=10, pady=2, sticky=(tk.W, tk.E))
        
        self.page_progress_label = ttk.Label(
            progress_frame, 
            text="0%",
            font=("Arial", 9)
        )
        self.page_progress_label.grid(row=0, column=2, pady=2)
        
        # Втори прогрес бар - бележки
        ttk.Label(progress_frame, text="Изтеглени бележки:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.receipt_label = ttk.Label(
            progress_frame, 
            text="0 бележки",
            font=("Arial", 9, "bold"),
            foreground="darkgreen"
        )
        self.receipt_label.grid(row=1, column=1, sticky=tk.W, padx=10, pady=2)
        
        progress_frame.columnconfigure(1, weight=1)
        
        # Рамка за логове
        log_frame = ttk.LabelFrame(self.root, text="Детайлни логове", padding="10")
        log_frame.grid(row=7, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, 
            height=20, 
            width=80,
            font=("Consolas", 9)
        )
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Конфигурация на grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(7, weight=1)
        
        # Стилове
        style = ttk.Style()
        try:
            style.configure("Accent.TButton", font=("Arial", 10, "bold"))
        except:
            pass
    
    def clear_start_date(self):
        """Изчиства началната дата"""
        self.use_period_var.set(False)
        self.log_message("✓ Изчистена начална дата")
    
    def clear_end_date(self):
        """Изчиства крайната дата"""
        self.use_period_var.set(False)
        self.log_message("✓ Изчистена крайна дата")
    
    def choose_directory(self):
        """Избира директория за съхранение"""
        directory = filedialog.askdirectory(
            title="Избери директория за съхранение на бележките",
            initialdir=self.output_dir
        )
        if directory:
            self.output_dir = directory
            self.dir_label.config(text=directory)
            self.log_message(f"✓ Избрана директория: {directory}")
    
    def load_config(self):
        """Зарежда конфигурацията от файла"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.output_dir = config.get('output_dir', str(Path.home() / "Documents"))
                    self.analysis_files = config.get('analysis_files', [])
        except Exception as e:
            print(f"Грешка при зареждане на конфигурация: {e}")
    
    def save_config(self):
        """Запазва конфигурацията във файла"""
        try:
            config = {}
            # Зареждаме съществуващата конфигурация ако има
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            # Обновяваме полетата
            config['output_dir'] = self.output_dir
            if self.analysis_files:
                config['analysis_files'] = self.analysis_files
            
            # Запазваме
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Грешка при запазване на конфигурация: {e}")
    
    def load_saved_analysis_file(self):
        """Зарежда запазените файлове за анализ в UI"""
        if self.analysis_files:
            # Проверка дали файловете съществуват
            existing_files = [f for f in self.analysis_files if os.path.exists(f)]
            self.analysis_files = existing_files
            
            if existing_files:
                count = len(existing_files)
                self.analysis_file_label.config(
                    text=f"Избрани {count} файла", 
                    foreground="blue"
                )
    
    def choose_analysis_files(self):
        """Избира множество файлове за анализ"""
        file_paths = filedialog.askopenfilenames(
            title="Избери файлове с касови бележки за анализ",
            initialdir=self.output_dir,
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if file_paths:
            self.analysis_files = list(file_paths)
            count = len(self.analysis_files)
            self.analysis_file_label.config(text=f"Избрани {count} файла", foreground="blue")
            
            self.log_message(f"✓ Избрани {count} файла за анализ")
            # Запазваме в конфига
            self.save_config()
    
    def choose_analysis_folder(self):
        """Избира папка и зарежда всички txt файлове от нея"""
        folder_path = filedialog.askdirectory(
            title="Избери папка с касови бележки",
            initialdir=self.output_dir
        )
        
        if folder_path:
            # Намиране на всички txt файлове в папката
            txt_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.txt')]
            
            if txt_files:
                self.analysis_files = txt_files
                count = len(txt_files)
                self.analysis_file_label.config(text=f"Избрани {count} файла от папка", foreground="blue")
                self.log_message(f"✓ Намерени {count} txt файла в папката")
                # Запазваме в конфига
                self.save_config()
            else:
                messagebox.showwarning(
                    "Внимание",
                    "Няма намерени txt файлове в избраната папка!"
                )
                self.log_message("⚠ Няма намерени txt файлове в папката")
    
    def choose_analysis_file(self):
        """Избира файл за анализ (запазва се за обратна съвместимост)"""
        file_path = filedialog.askopenfilename(
            title="Избери файл с касови бележки за анализ",
            initialdir=self.output_dir,
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if file_path:
            self.analysis_files = [file_path]
            # Показваме само името на файла, не целия път (за да се побере в интерфейса)
            file_name = os.path.basename(file_path)
            self.analysis_file_label.config(text=f"Избран 1 файл", foreground="blue")
            self.log_message(f"✓ Избран файл за анализ: {file_name}")
            # Запазваме в конфига
            self.save_config()
    
    def log_message(self, message):
        """Добавя съобщение в лог текста"""
        def _log():
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
        
        if threading.current_thread() != threading.main_thread():
            self.root.after(0, _log)
        else:
            _log()
    
    def update_status(self, message, color="black"):
        """Обновява статус лейбъла"""
        def _update():
            self.status_label.config(text=message, foreground=color)
        
        if threading.current_thread() != threading.main_thread():
            self.root.after(0, _update)
        else:
            _update()
    
    def update_progress(self, percent, page_num, total_pages, elapsed_time):
        """Обновява прогрес баровете и таймъра"""
        def _update():
            self.page_progress['value'] = percent
            self.page_progress_label.config(text=f"{int(percent)}% (стр. {page_num}/{total_pages})")
            
            receipt_count = len(self.downloader.receipts) if self.downloader else 0
            self.receipt_label.config(text=f"{receipt_count} бележки")
            
            # Обновяване на таймера
            hours = int(elapsed_time // 3600)
            minutes = int((elapsed_time % 3600) // 60)
            seconds = int(elapsed_time % 60)
            if hours > 0:
                time_str = f"{hours}ч {minutes}м {seconds}с"
            elif minutes > 0:
                time_str = f"{minutes}м {seconds}с"
            else:
                time_str = f"{seconds}с"
            self.timer_label.config(text=f"⏱ Време: {time_str}")
        
        if threading.current_thread() != threading.main_thread():
            self.root.after(0, _update)
        else:
            _update()
    
    def start_download(self):
        """Стартира изтеглянето"""
        if not os.path.exists(self.output_dir):
            messagebox.showerror("Грешка", "Избраната директория не съществува!")
            return
        
        # Получаване на датите от DateEntry виджетите
        start_date = None
        end_date = None
        
        if self.use_period_var.get():
            try:
                start_date = self.start_date_entry.get_date().strftime('%Y-%m-%d')
            except:
                start_date = None
            
            try:
                end_date = self.end_date_entry.get_date().strftime('%Y-%m-%d')
            except:
                end_date = None
            
            if start_date and end_date and start_date > end_date:
                messagebox.showerror("Грешка", "Началната дата не може да е след крайната!")
                return
        
        # Деактивиране на контроли
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.start_date_entry.config(state=tk.DISABLED)
        self.end_date_entry.config(state=tk.DISABLED)
        self.dir_button.config(state=tk.DISABLED)
        self.continue_button.config(state=tk.NORMAL)
        
        # Изчистване на логовете
        self.log_text.delete(1.0, tk.END)
        
        # Нулиране на прогрес баровете
        self.page_progress['value'] = 0
        self.page_progress_label.config(text="0%")
        self.receipt_label.config(text="0 бележки")
        self.timer_label.config(text="⏱ Време: 0с")
        
        self.update_status("⏳ Изчакване за влизане...", "orange")
        
        # Създаване на downloader
        self.downloader = LidlReceiptDownloader(
            self.output_dir,
            start_date=start_date,
            end_date=end_date,
            log_callback=self.log_message,
            progress_callback=self.update_progress
        )
        
        # Стартиране в отделна нишка
        self.download_thread = threading.Thread(target=self.run_download, daemon=True)
        self.download_thread.start()
    
    def run_download(self):
        """Изпълнява изтеглянето в отделна нишка"""
        try:
            # Създаване на нов event loop за тази нишка
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            # Изпълнение на изтеглянето
            loop.run_until_complete(self.downloader.download_all_receipts())
            
            # Запазване на файла
            if self.downloader.receipts and not self.downloader.is_cancelled:
                file_path = self.downloader.save_to_file()
                self.root.after(0, lambda: messagebox.showinfo(
                    "Успех", 
                    f"Успешно изтеглени {len(self.downloader.receipts)} бележки!\n\n"
                    f"Файл: {file_path}"
                ))
                self.update_status("✓ Завършено успешно", "green")
            elif self.downloader.is_cancelled:
                self.update_status("⚠ Прекъснато", "orange")
                if self.downloader.receipts:
                    file_path = self.downloader.save_to_file()
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Прекъснато", 
                        f"Процесът беше прекъснат.\n"
                        f"Запазени {len(self.downloader.receipts)} бележки.\n\n"
                        f"Файл: {file_path}"
                    ))
            else:
                self.update_status("⚠ Няма намерени бележки", "orange")
                self.root.after(0, lambda: messagebox.showwarning(
                    "Внимание", 
                    "Не са намерени касови бележки.\n\n"
                    "Възможни причини:\n"
                    "- Няма покупки в историята\n"
                    "- Проблем с влизането в акаунта\n"
                    "- Структурата на сайта е променена\n"
                    "- Всички бележки са извън избрания период"
                ))
                
        except Exception as e:
            self.log_message(f"\n❌ Грешка: {e}")
            self.update_status("❌ Грешка", "red")
            self.root.after(0, lambda: messagebox.showerror(
                "Грешка", 
                f"Възникна грешка при изтеглянето:\n\n{str(e)}"
            ))
        finally:
            # Активиране на контроли
            self.root.after(0, self.reset_ui)
    
    def continue_after_ready(self):
        """Продължава след като потребителя е готов"""
        if self.downloader:
            self.downloader.ready_to_start = True
            self.continue_button.config(state=tk.DISABLED)
            self.update_status("📥 Изтегляне...", "blue")
            self.log_message("\n✓ Стартиране на изтегляне на бележки...")
    
    def stop_download(self):
        """Прекъсва изтеглянето"""
        if self.downloader:
            self.downloader.is_cancelled = True
            self.log_message("\n⚠ Изпращане на сигнал за прекъсване...")
            self.stop_button.config(state=tk.DISABLED)
    
    def reset_ui(self):
        """Връща UI в начално състояние"""
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.continue_button.config(state=tk.DISABLED)
        self.start_date_entry.config(state=tk.NORMAL)
        self.end_date_entry.config(state=tk.NORMAL)
        self.dir_button.config(state=tk.NORMAL)
    
    def analyze_receipts(self):
        """Анализира касовите бележки и създава XLSX файл с история на цените"""
        # Проверка дали са избрани файлове
        if not self.analysis_files:
            # Ако няма избрани файлове, отваряме диалог
            file_paths = filedialog.askopenfilenames(
                title="Избери файлове с касови бележки за анализ",
                initialdir=self.output_dir,
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if not file_paths:
                return
            
            self.analysis_files = list(file_paths)
            count = len(self.analysis_files)
            self.analysis_file_label.config(text=f"Избрани {count} файла", foreground="blue")
        
        # Проверка дали файловете съществуват
        existing_files = [f for f in self.analysis_files if os.path.exists(f)]
        if not existing_files:
            messagebox.showerror(
                "Грешка", 
                f"Избраните файлове не съществуват!\n\nМоля изберете други файлове."
            )
            self.analysis_files = []
            self.analysis_file_label.config(text="Няма избрани файлове", foreground="gray")
            return
        
        # Актуализираме списъка със съществуващи файлове
        self.analysis_files = existing_files
        
        self.log_message(f"\n📊 Стартиране на анализ на {len(self.analysis_files)} файла:")
        for file_path in self.analysis_files:
            self.log_message(f"   • {os.path.basename(file_path)}")
        
        self.update_status("📊 Анализ...", "blue")
        
        try:
            # Парсване на всички файлове
            products_data = self.parse_receipts_files(self.analysis_files)
            
            if not products_data:
                messagebox.showwarning("Внимание", "Не са намерени артикули за анализ!")
                self.update_status("⚠ Няма данни", "orange")
                return
            
            # Филтриране на продукти, които се срещат повече от веднъж
            filtered_products = {
                product: dates_prices 
                for product, dates_prices in products_data.items() 
                if len(dates_prices) > 1
            }
            
            if not filtered_products:
                messagebox.showwarning(
                    "Внимание", 
                    "Не са намерени артикули, които се срещат повече от веднъж!"
                )
                self.update_status("⚠ Няма данни", "orange")
                return
            
            self.log_message(f"✓ Намерени {len(filtered_products)} артикула с повече от 1 покупка")
            self.log_message(f"  (Общо {len(products_data)} уникални артикула)")
            
            # Генериране на XLSX файл - използваме първия файл като база за името
            base_file = self.analysis_files[0]
            output_file = self.generate_xlsx(filtered_products, base_file)
            
            self.log_message(f"\n✓ XLSX файлът е създаден успешно!")
            self.log_message(f"  Файл: {output_file}")
            
            # Генериране на графика
            self.log_message(f"\n📊 Генериране на графика...")
            chart_file = self.generate_chart(output_file)
            
            if chart_file:
                self.log_message(f"✓ Графиката е създадена успешно!")
                self.log_message(f"  Файл: {chart_file}")
            
            self.update_status("✓ Анализ завършен", "green")
            
            messagebox.showinfo(
                "Успех", 
                f"Анализът завърши успешно!\n\n"
                f"Артикули с повече от 1 покупка: {len(filtered_products)}\n"
                f"Общо уникални артикули: {len(products_data)}\n\n"
                f"XLSX: {os.path.basename(output_file)}\n"
                f"Графика: {os.path.basename(chart_file) if chart_file else 'N/A'}"
            )
            
        except Exception as e:
            self.log_message(f"\n❌ Грешка при анализ: {e}")
            self.update_status("❌ Грешка при анализ", "red")
            messagebox.showerror("Грешка", f"Грешка при анализ:\n\n{str(e)}")
    
    def parse_receipts_files(self, file_paths):
        """Парсва множество файлове с бележки и извлича продукти с дати и цени"""
        products_data = defaultdict(dict)  # {product_name: {date: price}}
        
        total_receipts = 0
        
        for file_idx, file_path in enumerate(file_paths, 1):
            self.log_message(f"\n📄 Файл {file_idx}/{len(file_paths)}: {os.path.basename(file_path)}")
            
            try:
                file_products = self.parse_receipts_file(file_path)
                
                # Обединяване на данните от този файл с общите данни
                for product_name, dates_prices in file_products.items():
                    for date, price in dates_prices.items():
                        # Ако продуктът вече има цена за тази дата, използваме средната стойност
                        if date in products_data[product_name]:
                            # Средна стойност между двете цени
                            existing_price = products_data[product_name][date]
                            products_data[product_name][date] = (existing_price + price) / 2
                        else:
                            products_data[product_name][date] = price
                
                # Преброяване на бележките в този файл
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    receipts = content.split('БЕЛЕЖКА #')
                    file_receipt_count = len(receipts) - 1
                    total_receipts += file_receipt_count
                
            except Exception as e:
                self.log_message(f"  ⚠ Грешка при четене на файл: {e}")
                continue
        
        self.log_message(f"\n✓ Общо обработени: {total_receipts} бележки от {len(file_paths)} файла")
        self.log_message(f"✓ Намерени: {len(products_data)} уникални артикула")
        
        return products_data
    
    def parse_receipts_file(self, file_path):
        """Парсва файла с бележки и извлича продукти с дати и цени"""
        products_data = defaultdict(dict)  # {product_name: {date: price}}
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Разделяне на бележки
            receipts = content.split('БЕЛЕЖКА #')
            
            self.log_message(f"  ✓ Намерени {len(receipts)-1} бележки за парсинг...")
            
            for receipt_idx, receipt in enumerate(receipts[1:], 1):  # Прескачаме първия празен елемент
                # Извличане на дата - търсим различни формати
                date_match = None
                receipt_date_str = None
                
                # Формат 1: DD.MM.YYYY HH:MM:SS в края (например "13.01.2026 13:03:32")
                date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})\s+\d{2}:\d{2}:\d{2}', receipt)
                if date_match:
                    day, month, year = date_match.groups()
                    receipt_date_str = f"{year}-{month}-{day}"
                
                # Формат 2: YYYY.MM.DD HH:MM (например "2025.12.26 17:24")
                if not receipt_date_str:
                    date_match = re.search(r'(\d{4})\.(\d{2})\.(\d{2})\s+\d{2}:\d{2}', receipt)
                    if date_match:
                        year, month, day = date_match.groups()
                        receipt_date_str = f"{year}-{month}-{day}"
                
                # Формат 3: В заглавието "DD.месец" (например "13.януари")
                if not receipt_date_str:
                    months_bg = {
                        'януари': '01', 'февруари': '02', 'март': '03', 'април': '04',
                        'май': '05', 'юни': '06', 'юли': '07', 'август': '08',
                        'септември': '09', 'октомври': '10', 'ноември': '11', 'декември': '12'
                    }
                    for month_name, month_num in months_bg.items():
                        if month_name in receipt.lower():
                            day_match = re.search(r'(\d{1,2})\.' + month_name, receipt.lower())
                            if day_match:
                                day = day_match.group(1).zfill(2)
                                # Определяме годината - ако месецът е декември и сме в януари, значи е миналата година
                                year = '2025' if month_name == 'декември' else '2026'
                                receipt_date_str = f"{year}-{month_num}-{day}"
                                break
                
                if not receipt_date_str:
                    self.log_message(f"  ⚠ Пропусната бележка #{receipt_idx} - не може да се извлече дата")
                    continue
                
                try:
                    receipt_date = datetime.strptime(receipt_date_str, '%Y-%m-%d')
                except ValueError:
                    self.log_message(f"  ⚠ Пропусната бележка #{receipt_idx} - невалиден формат на дата: {receipt_date_str}")
                    continue
                
                # Определяне на конверсионен фактор и валута
                # Проверяваме дали е BGN, лв или Евро
                is_bgn = 'BGN' in receipt or '# лв' in receipt or 'лв  #' in receipt
                is_eur = 'Евро' in receipt or '# Евро #' in receipt or 'EUR' in receipt
                
                # Ако е преди 01.01.2026, трябва да конвертираме от BGN към EUR
                if receipt_date < datetime(2026, 1, 1):
                    # Стари бележки - винаги са в BGN, трябва да конвертираме
                    conversion_rate = 1.95583
                else:
                    # Нови бележки - ако са в BGN, конвертираме, ако са в EUR, оставяме
                    conversion_rate = 1.95583 if is_bgn else 1.0
                
                # Извличане на артикули и цени
                lines = receipt.split('\n')
                products_found = 0
                
                # Обработка на редовете - запазваме индекса за обработка на килограмови продукти
                for i, line in enumerate(lines):
                    # Прескачаме редове с маркери за купони и отстъпки
                    if any(marker in line for marker in ['#Lidl Plus купон', '#Акция', 'ОТСТЪПКИ', 
                                                          'МЕЖДИННА СУМА', 'ОБЩА СУМА', 'В БРОЙ',
                                                          'КРЕДИТНА/ДЕБИТНА', 'РЕСТО', '-----',
                                                          'Ти спести', '#Ном:', '#Z-отчет:', '#Каса:']):
                        continue
                    
                    # Шаблони за различни формати цени
                    # Формат 1: "ПРОДУКТ    ЦЕНА B" или "ПРОДУКТ    ЦЕНА лв"
                    price_pattern1 = r'^([А-ЯA-Z][А-ЯA-ZА-Яа-я\s\.\,\'\"\-\/\(\)0-9]+?)\s{2,}(\d+[\.,]\d{2})\s*[BDлв]*\s*$'
                    match = re.match(price_pattern1, line.strip())
                    
                    if match:
                        product_name = match.group(1).strip()
                        price_str = match.group(2).replace(',', '.')
                        
                        try:
                            price = float(price_str)
                        except ValueError:
                            continue
                        
                        # Прескачаме очевидни не-продукти
                        skip_keywords = ['ОБЩА', 'ОБЩО', 'ПЛАТЕНО', 'СУМА', 'TOTAL', 'PAID', 'НАЛИЧНОСТ', 
                                       'МЕЖДИННА', 'ОТСТЪПКИ', 'DISCOUNT', 'БАНКОВА', 'КАРТА',
                                       'ВАУЧЕР', 'VOUCHER', 'СДАЧА', 'CHANGE', 'РЕСТО', 'В БРОЙ',
                                       'ЗА ПЛАЩАНЕ', 'ПЛАЩАНЕ', 'FOR PAYMENT', 'PAYMENT',
                                       'Ном:', 'Z-отчет', 'Каса:', 'Касиер:', 'АРТИКУЛА', 'Копие',
                                       'ЕЛ. КУПОН', 'ЕЛ.КУПОН', 'КУПОН']
                        
                        if any(keyword in product_name.upper() for keyword in skip_keywords):
                            continue
                        
                        # Прескачаме твърде къси имена
                        if len(product_name) < 3:
                            continue
                        
                        # Прескачаме редове с количество (напр. "2,000 x 3,37")
                        if 'x' in product_name.lower() or 'х' in product_name.lower():
                            continue
                        
                        final_price = price
                        
                        # Проверяваме дали на предходния ред има шаблон "количество x единична_цена"
                        # Това се прилага за всички продукти (на КГ, на бройки и др.)
                        if i > 0:
                            prev_line = lines[i-1].strip()
                            # Формат: "количество x единична_цена" (напр. "3,000 x 3,37" или "0,890 x 2,55")
                            unit_price_pattern = r'(\d+[\.,]\d+)\s*[xх]\s*(\d+[\.,]\d{2})'
                            unit_match = re.search(unit_price_pattern, prev_line)
                            
                            if unit_match:
                                # Използваме единичната цена вместо крайната цена
                                unit_price_str = unit_match.group(2).replace(',', '.')
                                try:
                                    unit_price = float(unit_price_str)
                                    # Конвертиране на цена ако е нужно
                                    final_price = unit_price / conversion_rate
                                except ValueError:
                                    # Ако не може да се парсне, използваме оригиналната цена
                                    final_price = price / conversion_rate
                            else:
                                # Ако не намерим шаблона, използваме оригиналната цена
                                final_price = price / conversion_rate
                        else:
                            # Конвертиране на цена ако е нужно
                            final_price = price / conversion_rate
                        
                        # Съхраняване на данните
                        products_data[product_name][receipt_date_str] = final_price
                        products_found += 1
                
                if products_found > 0:
                    self.log_message(f"    ✓ Бележка #{receipt_idx} ({receipt_date_str}): {products_found} артикула")
            
            self.log_message(f"  ✓ От този файл: {len(products_data)} уникални артикула")
            return products_data
            
        except Exception as e:
            raise Exception(f"Грешка при четене на файла: {e}")
    
    def generate_xlsx(self, products_data, source_file):
        """Генерира XLSX файл с анализ на цените"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Грешка", 
                "Библиотеката 'openpyxl' не е инсталирана!\n\n"
                "Моля инсталирайте я с командата:\n"
                "pip install openpyxl"
            )
            raise ImportError("openpyxl is not installed")
        
        # Създаване на работна книга
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Price History"
        
        # Събиране на всички уникални дати и сортиране
        all_dates = set()
        for dates_prices in products_data.values():
            all_dates.update(dates_prices.keys())
        
        sorted_dates = sorted(all_dates)
        
        # Създаване на хедър
        ws['A1'] = "Артикул"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Добавяне на дати като колони
        for idx, date in enumerate(sorted_dates, start=2):
            col_letter = get_column_letter(idx)
            # Форматиране на датата за по-добра четливост (DD.MM.YYYY)
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d.%m.%Y')
            ws[f'{col_letter}1'] = formatted_date
            ws[f'{col_letter}1'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'{col_letter}1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Добавяне на данни за продуктите
        row_idx = 2
        for product_name in sorted(products_data.keys()):
            ws[f'A{row_idx}'] = product_name
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
            
            dates_prices = products_data[product_name]
            
            for col_idx, date in enumerate(sorted_dates, start=2):
                col_letter = get_column_letter(col_idx)
                
                if date in dates_prices:
                    price = dates_prices[date]
                    cell = ws[f'{col_letter}{row_idx}']
                    cell.value = price
                    cell.number_format = '[$€-407] #,##0.00'  # EUR формат
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            row_idx += 1
        
        # Настройка на ширина на колоните
        ws.column_dimensions['A'].width = 50
        for col_idx in range(2, len(sorted_dates) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # Добавяне на бордъри
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=len(sorted_dates)+1):
            for cell in row:
                cell.border = thin_border
        
        # Замръзване на първия ред и първата колона
        ws.freeze_panes = 'B2'
        
        # Генериране на име на файла
        base_name = os.path.splitext(source_file)[0]
        output_file = f"{base_name}_price_analysis.xlsx"
        
        # Запазване на файла
        wb.save(output_file)
        
        return output_file
    
    def generate_chart(self, xlsx_file):
        """
        Генерира графика на промяната на цените във времето от XLSX файл
        """
        try:
            import openpyxl
            from datetime import datetime
            
            # Четене на XLSX файла
            wb = openpyxl.load_workbook(xlsx_file)
            ws = wb.active
            
            # Извличане на дати от хедъра (ред 1, от колона 2 нататък)
            dates = []
            for col in range(2, ws.max_column + 1):
                date_str = ws.cell(row=1, column=col).value
                if date_str:
                    try:
                        date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                        dates.append(date_obj)
                    except:
                        continue
            
            if not dates:
                self.log_message("⚠ Не са намерени дати в XLSX файла")
                return None
            
            # Първо - преброяваме продуктите с повече от 5 цени
            products_with_enough_data = []
            
            for row_idx in range(2, ws.max_row + 1):
                product_name = ws.cell(row=row_idx, column=1).value
                if not product_name:
                    continue
                
                # Броене на попълнени цени за този продукт
                price_count = 0
                prices = []
                valid_dates = []
                
                for col_idx, date in enumerate(dates, start=2):
                    price_value = ws.cell(row=row_idx, column=col_idx).value
                    if price_value is not None:
                        price_count += 1
                        prices.append(float(price_value))
                        valid_dates.append(date)
                
                # Добавяме само ако има повече от 5 цени
                if price_count > 5:
                    products_with_enough_data.append({
                        'name': product_name,
                        'dates': valid_dates,
                        'prices': prices
                    })
            
            if not products_with_enough_data:
                self.log_message("⚠ Няма продукти с повече от 5 ценови записа")
                return None
            
            self.log_message(f"✓ Намерени {len(products_with_enough_data)} продукта с повече от 5 цени")
            
            # Генериране на интерактивна HTML графика с Plotly
            base_name = os.path.splitext(xlsx_file)[0]
            html_file = f"{base_name}_interactive_chart.html"
            
            # Създаване на интерактивна графика
            fig = go.Figure()
            
            # Добавяне на линии за всеки продукт
            for product_data in products_with_enough_data:
                product_name = product_data['name']
                valid_dates = product_data['dates']
                prices = product_data['prices']
                
                fig.add_trace(go.Scatter(
                    x=valid_dates,
                    y=prices,
                    mode='lines+markers',
                    name=product_name,
                    hovertemplate='<b>%{fullData.name}</b><br>' +
                                  'Дата: %{x|%d.%m.%Y}<br>' +
                                  'Цена: %{y:.2f} €<br>' +
                                  '<extra></extra>',
                    line=dict(width=2),
                    marker=dict(size=6)
                ))
            
            # Конфигурация на графиката
            fig.update_layout(
                title={
                    'text': f'Промяна на цените на продуктите във времето<br><sub>Продукти с повече от 5 записа: {len(products_with_enough_data)}</sub>',
                    'x': 0.5,
                    'xanchor': 'center',
                    'font': {'size': 20}
                },
                xaxis=dict(
                    title=dict(text='Дата', font=dict(size=14)),
                    tickformat='%d.%m.%Y',
                    gridcolor='lightgray'
                ),
                yaxis=dict(
                    title=dict(text='Цена (€)', font=dict(size=14)),
                    gridcolor='lightgray'
                ),
                hovermode='closest',
                template='plotly_white',
                height=800,
                showlegend=True,
                legend=dict(
                    orientation='v',
                    yanchor='top',
                    y=1,
                    xanchor='left',
                    x=1.02,
                    bgcolor='rgba(255, 255, 255, 0.9)',
                    bordercolor='lightgray',
                    borderwidth=1
                ),
                margin=dict(l=60, r=300, t=100, b=60)
            )
            
            # Добавяне на интерактивни контроли
            fig.update_xaxes(rangeslider_visible=True)
            
            # Изчисляване на топ 10 продукти с най-голяма промяна
            product_changes = []
            for product_data in products_with_enough_data:
                product_name = product_data['name']
                valid_dates = product_data['dates']
                prices = product_data['prices']
                
                if len(prices) >= 2:
                    first_price = prices[0]
                    last_price = prices[-1]
                    min_price = min(prices)
                    max_price = max(prices)
                    
                    # Намиране на датите на мин. и макс. цени
                    min_price_idx = prices.index(min_price)
                    max_price_idx = prices.index(max_price)
                    min_price_date = valid_dates[min_price_idx].strftime('%d.%m.%Y')
                    max_price_date = valid_dates[max_price_idx].strftime('%d.%m.%Y')
                    
                    # Процентна промяна: ((последна - първа) / първа) * 100
                    if first_price > 0:
                        percent_change = ((last_price - first_price) / first_price) * 100
                        product_changes.append({
                            'name': product_name,
                            'change_percent': percent_change,
                            'min_price': min_price,
                            'max_price': max_price,
                            'min_price_date': min_price_date,
                            'max_price_date': max_price_date,
                            'first_price': first_price,
                            'last_price': last_price
                        })
            
            # Сортиране по абсолютна стойност на промяната (най-голяма промяна първа)
            product_changes.sort(key=lambda x: abs(x['change_percent']), reverse=True)
            top_10_changes = product_changes[:10]
            
            # Генериране на HTML таблица
            table_html = '''
            <div class="table-container" style="margin-top: 30px; padding: 20px; background-color: white; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h2 style="color: #333; margin-bottom: 20px;">📊 Топ 10 продукти с най-голяма ценова промяна</h2>
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <thead>
                        <tr style="background-color: #007bff; color: white;">
                            <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">#</th>
                            <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">Продукт</th>
                            <th style="padding: 12px; text-align: center; border: 1px solid #ddd;">Промяна (%)</th>
                            <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Мин. цена (€)</th>
                            <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Макс. цена (€)</th>
                        </tr>
                    </thead>
                    <tbody>'''
            
            for idx, item in enumerate(top_10_changes, 1):
                # Цвят според промяната
                if item['change_percent'] > 0:
                    change_color = '#dc3545'  # червен за увеличение
                    arrow = '↑'
                else:
                    change_color = '#28a745'  # зелен за намаление
                    arrow = '↓'
                
                row_bg = '#f8f9fa' if idx % 2 == 0 else 'white'
                
                table_html += f'''
                        <tr style="background-color: {row_bg};">
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{idx}</td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{item['name']}</td>
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: {change_color};">
                                {arrow} {item['change_percent']:+.2f}%
                            </td>
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">
                                {item['min_price']:.2f}<br>
                                <small style="color: #666;">({item['min_price_date']})</small>
                            </td>
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">
                                {item['max_price']:.2f}<br>
                                <small style="color: #666;">({item['max_price_date']})</small>
                            </td>
                        </tr>'''
            
            table_html += '''
                    </tbody>
                </table>
                <p style="margin-top: 15px; color: #666; font-size: 12px;">
                    <strong>Забележка:</strong> Промяната е изчислена като процентна разлика между първата и последната дата на среща на продукта в анализа.
                    <span style="color: #dc3545;">↑ Увеличение</span> | <span style="color: #28a745;">↓ Намаление</span>
                </p>
            </div>'''
            
            # Генериране на втора таблица - топ 10 понижения
            price_decreases = [item for item in product_changes if item['change_percent'] < 0]
            price_decreases.sort(key=lambda x: x['change_percent'])  # Сортиране по най-голямо намаление (най-негативни)
            top_10_decreases = price_decreases[:10]
            
            if top_10_decreases:
                decrease_table_html = '''
            <div class="table-container" style="margin-top: 30px; padding: 20px; background-color: white; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h2 style="color: #333; margin-bottom: 20px;">📉 Топ 10 продукти с най-голямо понижение на цените</h2>
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <thead>
                        <tr style="background-color: #28a745; color: white;">
                            <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">#</th>
                            <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">Продукт</th>
                            <th style="padding: 12px; text-align: center; border: 1px solid #ddd;">Понижение (%)</th>
                            <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Мин. цена (€)</th>
                            <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Макс. цена (€)</th>
                        </tr>
                    </thead>
                    <tbody>'''
                
                for idx, item in enumerate(top_10_decreases, 1):
                    row_bg = '#f8f9fa' if idx % 2 == 0 else 'white'
                    
                    decrease_table_html += f'''
                        <tr style="background-color: {row_bg};">
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{idx}</td>
                            <td style="padding: 10px; border: 1px solid #ddd;">{item['name']}</td>
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: #28a745;">
                                ↓ {item['change_percent']:.2f}%
                            </td>
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">
                                {item['min_price']:.2f}<br>
                                <small style="color: #666;">({item['min_price_date']})</small>
                            </td>
                            <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">
                                {item['max_price']:.2f}<br>
                                <small style="color: #666;">({item['max_price_date']})</small>
                            </td>
                        </tr>'''
                
                decrease_table_html += '''
                    </tbody>
                </table>
                <p style="margin-top: 15px; color: #666; font-size: 12px;">
                    <strong>Забележка:</strong> Показани са само продуктите с намаление на цената между първата и последната дата на среща.
                </p>
            </div>'''
            else:
                decrease_table_html = ''
            
            # Запазване като HTML с добавени контроли за филтриране
            html_content = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Lidl - Интерактивна графика на цените</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 100%;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .controls {{
            margin-bottom: 20px;
            padding: 15px;
            background-color: #f8f9fa;
            border-radius: 5px;
        }}
        .control-group {{
            margin-bottom: 15px;
        }}
        label {{
            font-weight: bold;
            margin-right: 10px;
            display: inline-block;
            width: 150px;
        }}
        input[type="text"] {{
            padding: 8px;
            width: 300px;
            border: 1px solid #ddd;
            border-radius: 4px;
        }}
        button {{
            padding: 8px 20px;
            margin: 5px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
        }}
        .btn-primary {{
            background-color: #007bff;
            color: white;
        }}
        .btn-primary:hover {{
            background-color: #0056b3;
        }}
        .btn-secondary {{
            background-color: #6c757d;
            color: white;
        }}
        .btn-secondary:hover {{
            background-color: #545b62;
        }}
        .btn-success {{
            background-color: #28a745;
            color: white;
        }}
        .btn-success:hover {{
            background-color: #218838;
        }}
        .info {{
            margin-top: 10px;
            padding: 10px;
            background-color: #d1ecf1;
            border-left: 4px solid #0c5460;
            color: #0c5460;
        }}
        #chart {{
            margin-top: 20px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🛒 Lidl - Интерактивна графика на цените</h1>
        
        <div class="controls">
            <div class="control-group">
                <label>🔍 Търси продукт:</label>
                <input type="text" id="searchInput" placeholder="Въведете име на продукт...">
                <button class="btn-primary" onclick="filterProducts()">Филтрирай</button>
            </div>
            
            <div class="control-group">
                <button class="btn-success" onclick="showAll()">Покажи всички</button>
                <button class="btn-secondary" onclick="hideAll()">Скрий всички</button>
                <button class="btn-secondary" onclick="resetView()">Възстанови изглед</button>
            </div>
            
            <div class="info">
                <strong>💡 Съвети:</strong>
                <ul style="margin: 5px 0;">
                    <li>Кликнете на продукт в легендата за да го покажете/скриете</li>
                    <li>Използвайте мишката за приближаване (scroll) и местене (drag)</li>
                    <li>Използвайте филтъра за търсене на конкретни продукти</li>
                    <li>Двоен клик на легендата изолира един продукт</li>
                </ul>
            </div>
        </div>
        
        <div id="chart"></div>
        
        {table_html}
        
        {decrease_table_html}
    </div>
    
    <script>
        var plotData = {fig.to_json()};
        var layout = plotData.layout;
        var data = plotData.data;
        var config = {{
            responsive: true,
            displayModeBar: true,
            modeBarButtonsToAdd: ['drawopenpath', 'eraseshape'],
            toImageButtonOptions: {{
                format: 'png',
                filename: 'lidl_prices_chart',
                height: 1080,
                width: 1920,
                scale: 2
            }}
        }};
        
        // Запазване на оригиналната видимост
        var originalVisibility = data.map(trace => trace.visible);
        
        Plotly.newPlot('chart', data, layout, config);
        
        function filterProducts() {{
            var searchText = document.getElementById('searchInput').value.toLowerCase();
            
            if (!searchText) {{
                showAll();
                return;
            }}
            
            var update = {{}};
            update.visible = data.map(function(trace) {{
                return trace.name.toLowerCase().includes(searchText);
            }});
            
            Plotly.restyle('chart', update);
        }}
        
        function showAll() {{
            var update = {{}};
            update.visible = data.map(() => true);
            Plotly.restyle('chart', update);
            document.getElementById('searchInput').value = '';
        }}
        
        function hideAll() {{
            var update = {{}};
            update.visible = data.map(() => 'legendonly');
            Plotly.restyle('chart', update);
        }}
        
        function resetView() {{
            Plotly.relayout('chart', {{
                'xaxis.autorange': true,
                'yaxis.autorange': true
            }});
        }}
        
        // Добавяне на Enter key за търсене
        document.getElementById('searchInput').addEventListener('keypress', function(e) {{
            if (e.key === 'Enter') {{
                filterProducts();
            }}
        }});
    </script>
</body>
</html>'''
            
            with open(html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.log_message(f"✓ Интерактивна графика включва {len(products_with_enough_data)} продукта")
            self.log_message(f"✓ HTML файлът е запазен: {html_file}")
            
            # Опционално генериране и на PNG графика (статична версия)
            try:
                fig_height = max(8, min(20, 8 + len(products_with_enough_data) * 0.3))
                plt.style.use('seaborn-v0_8-darkgrid')
                fig_static, ax = plt.subplots(figsize=(16, fig_height))
                
                for product_data in products_with_enough_data:
                    product_name = product_data['name']
                    valid_dates = product_data['dates']
                    prices = product_data['prices']
                    short_name = product_name[:35] + '...' if len(product_name) > 35 else product_name
                    ax.plot(valid_dates, prices, marker='o', linewidth=2, markersize=5, label=short_name, alpha=0.8)
                
                ax.set_xlabel('Дата', fontsize=12, weight='bold')
                ax.set_ylabel('Цена (€)', fontsize=12, weight='bold')
                ax.set_title(f'Промяна на цените на продуктите във времето (продукти с повече от 5 записа: {len(products_with_enough_data)})', 
                            fontsize=14, weight='bold', pad=20)
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m.%Y'))
                ax.xaxis.set_major_locator(mdates.AutoDateLocator())
                plt.xticks(rotation=45, ha='right')
                
                if len(products_with_enough_data) <= 15:
                    ax.legend(loc='best', fontsize=8, framealpha=0.9, ncol=1)
                else:
                    ncols = min(3, (len(products_with_enough_data) + 9) // 10)
                    ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize=7, framealpha=0.9, ncol=ncols)
                
                ax.grid(True, alpha=0.3)
                plt.tight_layout()
                
                chart_file = f"{base_name}_chart.png"
                plt.savefig(chart_file, dpi=200, bbox_inches='tight')
                plt.close(fig_static)
                
                self.log_message(f"✓ Статична PNG графика също запазена: {chart_file}")
            except Exception as png_error:
                self.log_message(f"⚠ Статичната PNG графика не можа да се генерира: {png_error}")
            
            return html_file
            
        except Exception as e:
            self.log_message(f"❌ Грешка при генериране на графика: {e}")
            return None


def main():
    root = tk.Tk()
    app = LidlGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
